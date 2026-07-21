"""Ralph Loop 7-2: DARTLens 표시 layer 헬퍼 테스트(status_display / has_sparse_sheet).

원본 데이터 불변 · 표시 layer 변환만 검증(웹 렌더링 계층과 분리된 순수 헬퍼).
"""
import openpyxl

from src import ui_helpers as uih


def test_status_display_maps_raw_to_korean():
    assert uih.status_display("PASS") == "분석 완료"
    assert uih.status_display("PASS_WITH_WARNINGS") == "분석 완료 · 판정 제한"
    assert uih.status_display("FAIL") == "분석 실패"
    assert uih.status_display("INSUFFICIENT_PEERS") == "표본 제한"
    assert uih.status_display("NOT_COMPUTABLE") == "계산 불가"


def test_status_display_passthrough_and_none():
    assert uih.status_display("HELLO") == "HELLO"   # 매핑 없으면 원문 유지
    assert uih.status_display(None) == ""
    assert uih.status_display(" PASS ") == "분석 완료"  # 공백 허용


def _wb_with_sheets(path, sheet_names):
    wb = openpyxl.Workbook()
    wb.active.title = sheet_names[0]
    for s in sheet_names[1:]:
        wb.create_sheet(s)
    wb.save(str(path))
    return path


def test_has_sparse_sheet_true(tmp_path):
    p = _wb_with_sheets(tmp_path / "sparse.xlsx",
                        ["00_README", "08_계산불가_및_제외사유", "09_제한적_peer_비교"])
    assert uih.has_sparse_sheet(p) is True


def test_has_sparse_sheet_false(tmp_path):
    p = _wb_with_sheets(tmp_path / "nosparse.xlsx",
                        ["00_README", "02_수익성", "08_계산불가_및_제외사유"])
    assert uih.has_sparse_sheet(p) is False


def test_has_sparse_sheet_safe_on_bad_file(tmp_path):
    bad = tmp_path / "bad.xlsx"
    bad.write_bytes(b"not-an-xlsx")
    assert uih.has_sparse_sheet(bad) is False        # 예외 없이 False


# --------------------------------------------------------------------------
# Loop 7-2A: 최근 결과 로더 일반화 + run status 도출
# --------------------------------------------------------------------------
def test_list_report_candidates_detects_non_samsung_and_sorts_newest(tmp_path):
    names = [
        "삼성전자_산업대비_이상징후_리포트_2025_20260709_114446.xlsx",
        "현대자동차_산업대비_이상징후_리포트_2025_20260709_114448.xlsx",
        "대한항공_산업대비_이상징후_리포트_2025_20260709_114455.xlsx",
        "한화솔루션_산업대비_이상징후_리포트_2025_20260709_114459.xlsx",
    ]
    for nm in names:
        (tmp_path / nm).write_bytes(b"x")
    (tmp_path / "현대자동차_산업대비_이상징후_리포트_2025_99999999_999999.xlsx.tmp").write_bytes(b"x")  # 무시
    cands = uih.list_report_candidates(tmp_path)
    companies = [c["company"] for c in cands]
    assert {"현대자동차", "대한항공", "한화솔루션"} <= set(companies)   # 삼성 외도 후보
    assert cands[0]["timestamp"] == "20260709_114459"                # 최신 우선
    assert all(not c["filename"].endswith(".tmp") for c in cands)     # .tmp 제외
    assert "·" in uih.candidate_label(cands[0])


def test_list_report_candidates_empty_when_none(tmp_path):
    assert uih.list_report_candidates(tmp_path) == []


def test_find_debug_prefers_company_prefixed(tmp_path):
    rep = tmp_path / "현대자동차_산업대비_이상징후_리포트_2025_20260709_114448.xlsx"
    rep.write_bytes(b"x")
    dbg = tmp_path / "benchmark_debug_현대자동차_2025_20260709_114448.xlsx"   # multi-target 명명
    dbg.write_bytes(b"x")
    assert uih.find_debug_for_report(tmp_path, rep, 2025) == dbg


def test_find_debug_fallback_no_company_prefix(tmp_path):
    rep = tmp_path / "삼성전자_산업대비_이상징후_리포트_2025_20260705_110417.xlsx"
    rep.write_bytes(b"x")
    dbg = tmp_path / "benchmark_debug_2025_20260705_110417.xlsx"          # run_loop3b 명명
    dbg.write_bytes(b"x")
    assert uih.find_debug_for_report(tmp_path, rep, 2025) == dbg


def test_resolve_pair_for_report_matched_and_missing(tmp_path):
    rep = tmp_path / "대한항공_산업대비_이상징후_리포트_2025_20260709_114455.xlsx"
    rep.write_bytes(b"x")
    (tmp_path / "benchmark_debug_대한항공_2025_20260709_114455.xlsx").write_bytes(b"x")
    pair = uih.resolve_pair_for_report(tmp_path, rep, 2025)
    assert pair["status"] == "matched" and pair["pair_ok"] and pair["timestamp"] == "20260709_114455"

    rep2 = tmp_path / "현대자동차_산업대비_이상징후_리포트_2025_20260709_114448.xlsx"
    rep2.write_bytes(b"x")
    pair2 = uih.resolve_pair_for_report(tmp_path, rep2, 2025)
    assert pair2["status"] == "debug_missing" and pair2["debug"] is None


def _make_report(tmp_path, name, judgement="정상 범위", with_sparse=False):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "02_수익성"
    ws.cell(1, 1, "안내: HIGH=산업 대비 높음")                     # 병합 안내행
    for j, h in enumerate(["비율명", "판정", "benchmark_quality"], 1):
        ws.cell(2, j, h)
    ws.cell(3, 1, "영업이익률"); ws.cell(3, 2, judgement); ws.cell(3, 3, "STRONG")
    if with_sparse:
        wb.create_sheet("09_제한적_peer_비교")
    p = tmp_path / name
    wb.save(str(p))
    return p


def test_report_status_token_pass(tmp_path):
    p = _make_report(tmp_path, "삼성전자_산업대비_이상징후_리포트_2025_20260709_120000.xlsx", "정상 범위")
    assert uih.report_status_token(p) == "PASS"
    assert uih.status_display(uih.report_status_token(p)) == "분석 완료"


def test_report_status_token_warns_on_peer_shortage(tmp_path):
    p = _make_report(tmp_path, "현대자동차_산업대비_이상징후_리포트_2025_20260709_120000.xlsx", "peer 부족")
    assert uih.report_status_token(p) == "PASS_WITH_WARNINGS"
    assert uih.status_display(uih.report_status_token(p)) == "분석 완료 · 판정 제한"


def test_report_status_token_warns_on_sparse_sheet(tmp_path):
    p = _make_report(tmp_path, "대한항공_산업대비_이상징후_리포트_2025_20260709_120000.xlsx",
                     "정상 범위", with_sparse=True)
    assert uih.report_status_token(p) == "PASS_WITH_WARNINGS"   # sparse 시트 존재 → 판정 제한


def test_report_status_token_pass_when_only_not_computable(tmp_path):
    # 소수 '계산 불가'(NOT_COMPUTABLE)만 있고 peer 부족/sparse 없음 → runner PASS와 일관되게 PASS
    p = _make_report(tmp_path, "한화솔루션_산업대비_이상징후_리포트_2025_20260709_120000.xlsx", "계산 불가")
    assert uih.report_status_token(p) == "PASS"


def test_status_helpers_do_not_leak_key():
    secret = "cafebabe" * 5
    assert secret not in uih.mask_key(secret)          # key 값 미노출 유지
    assert secret not in uih.key_status_text(secret)
