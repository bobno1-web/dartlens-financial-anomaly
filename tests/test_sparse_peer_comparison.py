"""Ralph Loop 6: sparse-peer 직접 비교 단위/통합 테스트(라이브 API 없음).

검증 대상:
  1) peer 3개 → HIGH/LOW/NORMAL 억지 판정 없이 직접 비교 row 생성
  2) 직접 비교에 실제 회사명(값) 사용
  3) Peer 1/Peer 2 익명 컬럼 없음
  4) peer≥min_peers → 기존 benchmark 흐름 유지(sparse 미생성)
  5) peer 0개 → 비교 불가 note
  6) sparse target → 09_제한적_peer_비교 시트 생성 / 충분 peer target → 미생성
  7) 09 시트에 '통계적 benchmark가 아님' 주의 문구 존재
  8) 충분 peer target 기존 9시트 구조/컬럼 불변(회귀 방지)
"""
import re

import openpyxl

from src import excel_report as xr
from src import sparse_peer_comparison as sp


def _pair(ratio, group, target_value, label, peers, median=None):
    """(comparison_row, pool_detail) 최소 fixture. peers=[(corp_name, value), ...]"""
    inc = [{"corp_code": f"C{i}", "corp_name": nm, "value": v} for i, (nm, v) in enumerate(peers)]
    comp = {"ratio": ratio, "group": group, "target_value": target_value, "label": label}
    pd = {"ratio": ratio, "group": group, "included": inc, "stats": {"n_companies": len(inc), "median": median}}
    return comp, pd


def _split(pairs):
    return [c for c, _ in pairs], [p for _, p in pairs]


# --- 1) peer 3개 → 억지 판정 없음 ---
def test_sparse_three_peers_no_forced_label():
    comp, pd = _split([_pair("영업이익률", "수익성", 0.12, "INSUFFICIENT_PEERS",
                             [("기아", 0.08), ("HL만도", 0.04), ("현대모비스", 0.06)], median=0.06)])
    out = sp.build_sparse_peer_comparison(comp, pd, min_peers=5, target_name="테스트회사")
    assert len(out) == 1
    row = out[0]
    assert row["peer_count"] == 3
    assert row["is_statistical_benchmark"] is False
    assert row["label"] not in ("HIGH", "LOW", "NORMAL")   # 부족 표본으로 통계 판정 생성 안 함
    assert row["sample_status"] == sp.STATUS_LIMITED


# --- 2) 실제 회사명(값) 사용 ---
def test_sparse_peer_values_use_real_company_names():
    comp, pd = _split([_pair("영업이익률", "수익성", 0.12, "INSUFFICIENT_PEERS",
                             [("기아", 0.08), ("HL만도", 0.04), ("현대모비스", 0.06)], median=0.06)])
    out = sp.build_sparse_peer_comparison(comp, pd, min_peers=5)
    names = [p["corp_name"] for p in out[0]["peer_company_values"]]
    assert set(names) == {"기아", "HL만도", "현대모비스"}
    vals = [p["value"] for p in out[0]["peer_company_values"]]
    assert vals == sorted(vals, reverse=True)               # 값 내림차순
    assert all(p["value"] is not None for p in out[0]["peer_company_values"])


# --- 3) Peer 1/Peer 2 익명 표기 없음 ---
def test_no_anonymous_peer_labels():
    comp, pd = _split([_pair("영업이익률", "수익성", 0.12, "INSUFFICIENT_PEERS",
                             [("기아", 0.08), ("HL만도", 0.04)], median=0.06)])
    out = sp.build_sparse_peer_comparison(comp, pd, min_peers=5)
    names = [p["corp_name"] for p in out[0]["peer_company_values"]]
    assert not any(re.search(r"(?i)peer\s*\d", n) for n in names)


# --- 4) peer≥min_peers → benchmark 유지(sparse 미생성) ---
def test_sufficient_peers_excluded_from_sparse():
    peers5 = [("A", 0.05), ("B", 0.06), ("C", 0.07), ("D", 0.08), ("E", 0.09)]
    comp, pd = _split([_pair("순이익률", "수익성", 0.10, "NORMAL", peers5, median=0.07)])
    out = sp.build_sparse_peer_comparison(comp, pd, min_peers=5)
    assert out == []     # 5개(=min_peers) → benchmark 성립, 직접비교 미생성


# --- 5) peer 0개 → 비교 불가 ---
def test_zero_peers_note():
    comp, pd = _split([_pair("매입채무비율", "운전자본", 0.03, "INSUFFICIENT_PEERS", [], median=None)])
    out = sp.build_sparse_peer_comparison(comp, pd, min_peers=5)
    assert len(out) == 1 and out[0]["peer_count"] == 0
    assert out[0]["sample_status"] == sp.STATUS_NONE
    assert out[0]["target_rank_among_peers"] is None
    assert out[0]["peer_company_values"] == []


def test_single_peer_status_and_rank():
    comp, pd = _split([_pair("ROA", "수익성", 0.05, "INSUFFICIENT_PEERS", [("기아", 0.04)], median=0.04)])
    out = sp.build_sparse_peer_comparison(comp, pd, min_peers=5)
    assert out[0]["sample_status"] == sp.STATUS_SINGLE and out[0]["peer_count"] == 1
    assert out[0]["target_rank_among_peers"] == "1/2 (값 내림차순)"   # target 0.05 > peer 0.04


def test_target_not_computable_keeps_peer_values_no_rank():
    """target 비율이 계산불가면 peer 값은 보여주되 순위는 보류(참고 비교 유지)."""
    comp, pd = _split([_pair("매입채무비율", "운전자본", None, "NOT_COMPUTABLE",
                             [("기아", 0.03), ("현대모비스", 0.05)], median=0.04)])
    out = sp.build_sparse_peer_comparison(comp, pd, min_peers=5)
    assert out[0]["target_value"] is None
    assert out[0]["target_rank_among_peers"] is None
    assert len(out[0]["peer_company_values"]) == 2


# --- 6~8) Excel 통합 ---
def _build_report(tmp_path, sparse_rows, corp="현대자동차(주)"):
    target = {"corp_name": corp, "stock_code": "005380", "corp_code": "0",
              "induty_code": "301", "effective_prefix": "301"}
    meta = {"bsns_year": 2025, "iqr_fence_k": 1.5, "min_peers": 5}
    return xr.build_final_report_workbook(
        tmp_path / "f.xlsx", target=target, target_cfs_rows=[], peers=[], peer_rows=[],
        comparison_rows=[], excluded_summary=[], meta=meta, sparse_comparison=sparse_rows)


def test_sparse_sheet_created_with_real_name_columns(tmp_path):
    comp, pd = _split([_pair("영업이익률", "수익성", 0.12, "INSUFFICIENT_PEERS",
                             [("기아", 0.08), ("HL만도", 0.04), ("현대모비스", 0.06)], median=0.06)])
    sparse = sp.build_sparse_peer_comparison(comp, pd, min_peers=5, target_name="현대자동차(주)")
    p = _build_report(tmp_path, sparse)
    wb = openpyxl.load_workbook(str(p), read_only=True)
    names = wb.sheetnames
    rows = list(wb["09_제한적_peer_비교"].iter_rows(values_only=True))
    wb.close()
    assert "09_제한적_peer_비교" in names
    header = rows[1]                                        # row0=주의문구(merge), row1=헤더
    assert "기아 값" in header and "HL만도 값" in header and "현대모비스 값" in header
    assert not any(h and re.search(r"(?i)peer\s*\d", str(h)) for h in header)   # 익명 컬럼 없음


def test_sufficient_peer_target_has_no_sparse_sheet(tmp_path):
    p = _build_report(tmp_path, None)                      # sparse 없음
    wb = openpyxl.load_workbook(str(p), read_only=True)
    names = wb.sheetnames
    wb.close()
    assert "09_제한적_peer_비교" not in names
    assert "08_계산불가_및_제외사유" in names               # 기존 구조 유지


def test_sparse_sheet_has_non_statistical_warning(tmp_path):
    comp, pd = _split([_pair("영업이익률", "수익성", 0.12, "INSUFFICIENT_PEERS",
                             [("기아", 0.08), ("HL만도", 0.04)], median=0.06)])
    sparse = sp.build_sparse_peer_comparison(comp, pd, min_peers=5)
    p = _build_report(tmp_path, sparse)
    wb = openpyxl.load_workbook(str(p), read_only=True)
    txt = " ".join(str(c) for row in wb["09_제한적_peer_비교"].iter_rows(values_only=True)
                   for c in row if c is not None)
    wb.close()
    assert "통계적 benchmark가 아" in txt and "min_peers" in txt


def test_backward_compat_default_no_sparse_sheet(tmp_path):
    """sparse_comparison 미전달 시 기존 9시트 구조/이름 불변(회귀 방지)."""
    target = {"corp_name": "테스트회사(주)", "stock_code": "000000", "corp_code": "0",
              "induty_code": "999", "effective_prefix": "999"}
    meta = {"bsns_year": 2025, "iqr_fence_k": 1.5, "min_peers": 5}
    p = xr.build_final_report_workbook(
        tmp_path / "g.xlsx", target=target, target_cfs_rows=[], peers=[], peer_rows=[],
        comparison_rows=[], excluded_summary=[], meta=meta)
    wb = openpyxl.load_workbook(str(p), read_only=True)
    names = wb.sheetnames
    wb.close()
    # Loop 19: 첫 시트로 '00_한눈에보기' 요약 신설(표시 layer). 나머지 구조/이름은 불변.
    assert names == ["00_한눈에보기", "00_README", "01_테스트회사_연결재무제표", "02_수익성",
                     "03_안정성_재무구조", "04_운전자본_계정리스크", "05_회전율", "06_Peer_List",
                     "07_Methodology", "08_계산불가_및_제외사유"]


def test_runner_summary_has_sparse_columns():
    from src import multi_target_runner as m
    blank = m._blank_row("005380", 2025, "hint")
    for k in ("sparse_peer_comparison_available", "sparse_peer_ratio_count",
              "sparse_peer_company_count_min", "sparse_peer_company_count_max", "sparse_peer_note"):
        assert k in blank
    keys = [k for k, _h in m.SUMMARY_COLUMNS]
    assert "sparse_peer_comparison_available" in keys and "sparse_peer_ratio_count" in keys
