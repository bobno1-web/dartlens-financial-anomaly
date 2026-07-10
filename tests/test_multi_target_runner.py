"""Ralph Loop 5/5-A: multi_target_runner + 시트명 helper 단위 테스트(라이브 API 없음)."""
from src import multi_target_runner as m
from src.excel_report import safe_sheet_name


def _cmp(label, quality):
    return {"label": label, "benchmark_quality": quality}


def test_safe_filename_strips_and_sanitizes():
    assert m.safe_filename("현대자동차(주)") == "현대자동차"
    assert m.safe_filename("CJ제일제당(주)") == "CJ제일제당"
    assert m.safe_filename("삼성전자(주)") == "삼성전자"
    assert m.safe_filename("a/b:c") == "a_b_c"             # 파일시스템 금지문자 → _
    assert m.safe_filename("  한화 솔루션 ") == "한화솔루션"   # 공백 제거
    assert m.safe_filename("") == "target"                # 빈 값 fallback


def test_label_counts():
    rows = [_cmp("NORMAL", "STRONG")] * 11 + [_cmp("HIGH", "STRONG")] + \
           [_cmp("INSUFFICIENT_PEERS", "NOT_AVAILABLE")] * 2 + [_cmp("NOT_COMPUTABLE", "NOT_AVAILABLE")]
    lc = m._label_counts(rows)
    assert lc["NORMAL"] == 11 and lc["HIGH"] == 1 and lc["INSUFFICIENT_PEERS"] == 2
    assert lc["NOT_COMPUTABLE"] == 1 and lc["LOW"] == 0


def test_quality_counts():
    rows = [_cmp("NORMAL", "STRONG")] * 12 + [_cmp("NORMAL", "WEAK")] * 2 + [_cmp("NORMAL", "LIMITED")]
    qc = m._quality_counts(rows)
    assert qc["STRONG"] == 12 and qc["WEAK"] == 2 and qc["LIMITED"] == 1 and qc["NOT_AVAILABLE"] == 0


def test_summary_columns_all_present_in_blank_row():
    blank = m._blank_row("005930", 2025, "hint")
    missing = [k for k, _h in m.SUMMARY_COLUMNS if k not in blank]
    assert missing == [], f"summary 컬럼이 row에 없음: {missing}"


def test_target_lists_have_stock_and_name():
    for t in m.DEFAULT_TARGETS + m.BACKUP_TARGETS:
        assert t["stock"].isdigit() and len(t["stock"]) == 6
        assert t["name"] and t["industry_hint"]
    # baseline은 삼성전자 005930
    assert any(t["stock"] == "005930" for t in m.DEFAULT_TARGETS)


# --- Loop 5-A: Excel 시트명 동적화 ---
def test_sheet_name_samsung_backward_compat():
    # 삼성전자는 기존 하드코딩 시트명과 완전히 동일해야 함(verify/기존 산출물 호환)
    assert safe_sheet_name("01_", "삼성전자(주)", "_연결재무제표") == "01_삼성전자_연결재무제표"
    assert safe_sheet_name("03_", "삼성전자(주)", "_comparison") == "03_삼성전자_comparison"


def test_sheet_name_dynamic_per_company():
    cases = {"CJ제일제당(주)": "01_CJ제일제당_연결재무제표",
             "한화솔루션(주)": "01_한화솔루션_연결재무제표",
             "현대자동차(주)": "01_현대자동차_연결재무제표",
             "대한항공(주)": "01_대한항공_연결재무제표"}
    for corp, exp in cases.items():
        got = safe_sheet_name("01_", corp, "_연결재무제표")
        assert got == exp
        assert "삼성전자" not in got          # 타 회사에 삼성전자 고정 금지


def test_sheet_name_limit_and_forbidden():
    n = safe_sheet_name("01_", "A/B:C[x]*D?", "_연결재무제표")
    assert not any(ch in n for ch in "\\/?*[]:")   # Excel 금지문자 제거
    long = safe_sheet_name("01_", "가" * 50, "_연결재무제표")
    assert len(long) <= 31                          # 31자 제한
    assert safe_sheet_name("01_", "", "_연결재무제표") == "01_대상_연결재무제표"  # 빈 이름 fallback


def test_final_report_01_sheet_name_dynamic_integration(tmp_path):
    """build_final_report_workbook이 실제로 corp_name 기반 01 시트명을 쓰는지(라이브 API 없이)."""
    import openpyxl
    from src import excel_report as xr
    target = {"corp_name": "테스트회사(주)", "stock_code": "000000", "corp_code": "00000000",
              "induty_code": "999", "effective_prefix": "999"}
    meta = {"bsns_year": 2025, "iqr_fence_k": 1.5, "min_peers": 5}
    p = xr.build_final_report_workbook(
        tmp_path / "t.xlsx", target=target, target_cfs_rows=[], peers=[], peer_rows=[],
        comparison_rows=[], excluded_summary=[], meta=meta)
    wb = openpyxl.load_workbook(str(p), read_only=True)
    names = wb.sheetnames
    wb.close()
    assert "01_테스트회사_연결재무제표" in names
    assert "01_삼성전자_연결재무제표" not in names


def _sheet_text(path, sheet):
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True)
    txt = " ".join(str(c) for row in wb[sheet].iter_rows(values_only=True) for c in row if c is not None)
    wb.close()
    return txt


def test_final_ratio_columns_generalized():
    from src import excel_report as xr
    assert "삼성전자 값" not in xr.FINAL_RATIO_COLUMNS   # 컬럼 고정 금지
    assert "대상회사 값" in xr.FINAL_RATIO_COLUMNS


def test_final_report_no_samsung_in_userfacing_text(tmp_path):
    """비삼성 target final report의 사용자 표시 텍스트/컬럼에 '삼성전자'가 남지 않는지."""
    from src import excel_report as xr
    target = {"corp_name": "테스트회사(주)", "stock_code": "000000", "corp_code": "00000000",
              "induty_code": "999", "effective_prefix": "999"}
    meta = {"bsns_year": 2025, "iqr_fence_k": 1.5, "min_peers": 5}
    p = xr.build_final_report_workbook(
        tmp_path / "f.xlsx", target=target, target_cfs_rows=[], peers=[], peer_rows=[],
        comparison_rows=[], excluded_summary=[], meta=meta)
    # 02 헤더는 '대상회사 값'(삼성전자 값 아님)
    hdr = _sheet_text(p, "02_수익성")
    assert "대상회사 값" in hdr and "삼성전자 값" not in hdr
    # README/Methodology 사실 설명문에 '삼성전자' 없음
    for sh in ("00_README", "07_Methodology"):
        assert "삼성전자" not in _sheet_text(p, sh), f"{sh}에 삼성전자 잔존"


def test_debug_workbook_no_samsung_in_userfacing_text(tmp_path):
    """비삼성 target benchmark_debug의 03 헤더/검증안내에 '삼성전자'가 남지 않는지."""
    from src import excel_report as xr
    target = {"corp_name": "테스트회사(주)", "stock_code": "000000", "corp_code": "00000000",
              "induty_code": "999", "effective_prefix": "999"}
    meta = {"bsns_year": 2025, "iqr_fence_k": 1.5, "min_peers": 5}
    p = xr.build_benchmark_debug_workbook(
        tmp_path / "d.xlsx", target=target, comparison_rows=[], pool_details=[],
        ratio_rows_all=[], target_trace=[], meta=meta)
    import openpyxl
    wb = openpyxl.load_workbook(str(p), read_only=True)
    names = wb.sheetnames
    wb.close()
    assert "03_테스트회사_comparison" in names and "03_삼성전자_comparison" not in names
    assert "삼성전자" not in _sheet_text(p, "00_검증안내")
    assert "대상회사값" in _sheet_text(p, "03_테스트회사_comparison")


# --- Loop 5-D: 잔존 삼성 표시(08 대상명 who, debug 04 basis "삼성값") 제거 ---
def test_excluded_summary_target_name_is_dynamic():
    from src.pipeline import _excluded_summary
    ds = {"target": {"corp_name": "테스트회사(주)"}, "excluded": []}
    comp = [{"label": "INSUFFICIENT_PEERS", "ratio": "영업이익률", "reason": "peer 부족"},
            {"label": "NORMAL", "ratio": "순이익률", "reason": ""}]
    out = _excluded_summary(comp, ds, [])
    rows = [r for r in out if r["kind"] == "비율 판정 불가/부족"]
    assert rows and all(r["who"] == "테스트회사(주)" for r in rows)   # 대상명 동적
    assert all(r["who"] != "삼성전자" for r in rows)                  # 삼성 고정 아님
    assert len(rows) == 1                                             # row 수 불변(라벨만 카운트)


def test_excluded_summary_failclose_when_no_corp_name():
    from src.pipeline import _excluded_summary
    ds = {"target": {}, "excluded": []}     # corp_name 없음 → fail-close
    out = _excluded_summary([{"label": "NOT_COMPUTABLE", "ratio": "매입채무비율", "reason": "missing_account"}], ds, [])
    assert out and out[0]["who"] == "대상회사"


def test_debug_04_basis_label_generalized(tmp_path):
    """debug 04_label_reason_상세 산정근거가 '대상값='(삼성값= 아님)인지 — 실데이터."""
    import openpyxl
    from src import excel_report as xr
    stats = {"n_companies": 3, "mean": 0.1, "winsorized_mean": 0.1, "median": 0.05, "p25": 0.02,
             "p75": 0.08, "iqr": 0.06, "mad": 0.03, "std": 0.04, "min": 0.01, "max": 0.2}
    pd = {"ratio": "영업이익률", "group": "수익성",
          "included": [{"corp_code": "P1", "corp_name": "p1", "value": 0.05}],
          "excluded": [{"corp_code": "T", "corp_name": "t", "reason": "target_leave_one_out", "is_target": True}],
          "stats": stats, "benchmark_quality": "WEAK", "quality_reason": "r",
          "upper_fence": 0.16, "lower_fence": -0.07, "cfs_success": 3}
    comp = [{"ratio": "영업이익률", "group": "수익성", "target_value": 0.13, "label": "NORMAL",
             "reason": "정상", "audit_comment": "코멘트", "robust_z": 1.2, "percentile": 90.0,
             "deviation_rate": 4.4, "deviation_pp_display": 10.6, "deviation_reason": "",
             "benchmark_quality": "WEAK", "interpret_note": "비고", "stats": stats}]
    target = {"corp_name": "테스트회사(주)", "stock_code": "000000", "corp_code": "0", "induty_code": "999"}
    p = xr.build_benchmark_debug_workbook(
        tmp_path / "d.xlsx", target=target, comparison_rows=comp, pool_details=[pd],
        ratio_rows_all=[], target_trace=[], meta={"bsns_year": 2025, "iqr_fence_k": 1.5, "min_peers": 5})
    txt = _sheet_text(p, "04_label_reason_상세")
    assert "대상값=" in txt and "삼성값=" not in txt
