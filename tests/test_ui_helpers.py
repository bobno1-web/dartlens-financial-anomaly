"""Ralph Loop 4 UI helper 단위 테스트 (웹 UI E2E 아님, 순수 헬퍼 단위).

Loop 11: 삼성 전용 legacy 탐지 함수(find_latest_report/find_latest_debug/resolve_report_pair)
제거에 따라 해당 9개 테스트를 제거했다. 동등한 최신탐지·pair 매칭 커버리지는 회사 일반 함수
(list_report_candidates / resolve_pair_for_report / find_debug_for_report) 기준으로
test_ui_dartlens.py 에 있다.
"""
import openpyxl

from src import ui_helpers as uih


# --- 합성 최종 리포트 xlsx (실제 구조 모사: 상단 안내행 + 헤더행) ---
def _make_final(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "02_수익성"
    ws.cell(1, 1, "안내: HIGH=산업 대비 높음 / NORMAL=정상 범위")  # 병합 안내행(row0)
    headers = ["비율명", "그룹", "산식", "삼성전자 값", "산업 중앙값", "중앙값 대비 차이(%p·값)",
               "robust_z", "percentile", "판정", "benchmark_quality",
               "peer 후보 수", "CFS 성공 peer 수", "CFS 실패 peer 수"]
    for j, h in enumerate(headers, 1):
        ws.cell(2, j, h)
    r1 = ["영업이익률", "수익성", "영업이익/매출액", 0.13, 0.02, 10.6, 1.2, 90.0, "정상 범위", "STRONG", 60, 51, 9]
    r2 = ["매출채권비율", "운전자본", "매출채권/매출액", 0.05, 0.04, 1.0, 0.3, 55.0, "정상 범위", "WEAK", 60, 51, 9]
    for i, row in enumerate((r1, r2), start=3):
        for j, v in enumerate(row, 1):
            ws.cell(i, j, v)

    ws2 = wb.create_sheet("06_Peer_List")
    ws2.cell(1, 1, "안내")
    for j, h in enumerate(["기업코드", "기업명", "대상여부"], 1):
        ws2.cell(2, j, h)
    for i, row in enumerate((["00126380", "테스트전자", "대상"], ["00111111", "피어사", "peer"]), start=3):
        for j, v in enumerate(row, 1):
            ws2.cell(i, j, v)

    ws3 = wb.create_sheet("08_계산불가_및_제외사유")
    for j, h in enumerate(["구분", "대상", "항목", "사유"], 1):
        ws3.cell(1, j, h)
    for j, v in enumerate(["CFS 사용불가(peer)", "피어사", "00111111", "status=013"], 1):
        ws3.cell(2, j, v)

    path = tmp_path / "삼성전자_산업대비_이상징후_리포트_2025_20260705_120000.xlsx"
    wb.save(str(path))
    return path


def test_sheet_to_df_detects_header_below_note(tmp_path):
    p = _make_final(tmp_path)
    df = uih.ratio_sheet_df(p, "02_수익성")
    assert "비율명" in df.columns and "판정" in df.columns
    assert set(df["비율명"]) == {"영업이익률", "매출채권비율"}


def test_extract_summary(tmp_path):
    p = _make_final(tmp_path)
    s = uih.extract_summary(p)
    assert s["company"] == "테스트전자"          # 06_Peer_List 대상행
    assert s["year"] == "2025"                    # 파일명
    assert s["peer_candidates"] == 60 and s["cfs_success"] == 51 and s["cfs_fail"] == 9
    assert s["computable_count"] == 2 and s["total_ratios"] == 2
    assert s["label_counts"].get("정상 범위") == 2


def test_build_interpretation_reads_data(tmp_path):
    p = _make_final(tmp_path)
    lines = uih.build_interpretation(p)
    joined = " ".join(lines)
    assert "정상 범위" in joined                  # 전부 NORMAL 문장
    assert "상위권" in joined                      # 영업이익률 percentile 90
    assert "WEAK" in joined or "LIMITED" in joined  # benchmark_quality 제한 안내


def test_build_interpretation_safe_default_when_unreadable(tmp_path):
    bad = tmp_path / "삼성전자_산업대비_이상징후_리포트_2025_20260705_120000.xlsx"
    bad.write_bytes(b"not-an-xlsx")
    lines = uih.build_interpretation(bad)
    assert isinstance(lines, list) and len(lines) >= 1  # 예외 없이 안전 기본 안내


def test_prepare_download_bytes(tmp_path):
    p = _make_final(tmp_path)
    name, data = uih.prepare_download(p)
    assert name == p.name and isinstance(data, (bytes, bytearray)) and len(data) > 0


# --- API key 안전 처리: key 값을 절대 반환/노출하지 않음 ---
def test_mask_key_never_leaks():
    # 40자리 hex(실 키 형태). 소스에 40-hex 리터럴을 직접 두지 않도록 조합해 생성
    # (secret-scanner/ pre-commit 훅 오탐 방지 — 런타임 값은 동일).
    secret = "0123456789abcdef" * 2 + "01234567"
    assert secret not in uih.mask_key(secret)
    assert uih.mask_key("") == "미설정"
    assert uih.mask_key(secret) == "설정됨(****)"


def test_key_status_text_does_not_reveal_key():
    secret = "deadbeef" * 5
    txt = uih.key_status_text(secret)
    assert secret not in txt and "세션" in txt
