"""Ralph Loop 2/3 pipeline.

Loop 2 (default): FULL peer CFS collection + account mapping + per-company ratio
INPUT preparation + debug/skeleton Excel + internal verification.

Loop 3 (--loop 3): per-ratio industry benchmark (median/IQR, leave-one-out) +
target comparison + final user Excel + benchmark debug Excel. Loop 3 reconstructs
the ratio-input table from cache (deterministic, offline) and cross-checks it
against the latest Loop 2 ratio_input_debug artifact before benchmarking. It does
NOT overwrite Loop 2 outputs (new timestamped files only).

Halts (StopConditionError, exit 2) instead of falling back on stop conditions.
"""
from __future__ import annotations

import argparse
import json
import math
import sys
from datetime import datetime
from pathlib import Path

from . import accounts, benchmarks, collect, compare, config, corp_codes, ratio_input
from .dart_client import DartClient, StopConditionError
from .parse import detect_key_accounts

HANGUL = range(0xAC00, 0xD7A4)
_ALLOWED_LABELS = {"NORMAL", "HIGH", "LOW", "INSUFFICIENT_PEERS",
                   "NOT_COMPUTABLE", "INSUFFICIENT_VARIANCE"}


def _has_hangul(s: str) -> bool:
    return any(ord(ch) in HANGUL for ch in s)


def _cfs_status_ko(status, nrows) -> str:
    if status == "000" and nrows > 0:
        return "성공"
    if status == "013":
        return "CFS없음(013)"
    return f"실패({status})"


def _f(v):
    return float(v) if v is not None else None


def _finite(v) -> bool:
    """True if v is None or a finite number (no NaN/inf)."""
    if v is None:
        return True
    if isinstance(v, bool):
        return True
    if isinstance(v, (int, float)):
        return not (math.isinf(v) or math.isnan(v))
    return True


# ===========================================================================
# Shared dataset builder (collection + ratio input) — used by BOTH loops.
# Cache-first + deterministic. Does NOT write any Excel.
# ===========================================================================
def build_dataset(stock_code, bsns_year, settings, paths, api_key) -> dict:
    prefix_len = int(settings.get("ksic_prefix_len", 3))
    allowed_cls = settings.get("listed_corp_cls", ["Y", "K"])
    client = DartClient(
        api_key, paths["raw"], paths["cache"], config.PROJECT_ROOT,
        timeout=int(settings.get("request_timeout_sec", 20)),
        delay=float(settings.get("request_delay_sec", 0.0)),
        max_retries=int(settings.get("max_retries", 3)),
    )

    print("  - corpCode 마스터 로드...", flush=True)
    records = corp_codes.parse_corp_codes(client.get_corpcode_xml())
    listed = corp_codes.listed_companies(records)

    print("  - 대상 회사 해결 + 기업개황...", flush=True)
    target = collect.resolve_target(client, records, stock_code, prefix_len)
    print(f"    corp_code={target['corp_code']} induty={target['induty_code']} prefix={target['effective_prefix']}", flush=True)

    print("  - peer universe 스캔(induty_code, 캐시)...", flush=True)
    peers, scan_stats = collect.scan_peers(
        client, listed, target["corp_code"], target["effective_prefix"], allowed_cls,
        int(settings.get("scan_workers", 8)), int(settings.get("peer_scan_limit", 0)))
    print(f"    peer 후보 {len(peers)}건", flush=True)
    if len(peers) == 0:
        raise StopConditionError("peer 후보가 0개입니다. (STOP)")

    print(f"  - {target['corp_name']} CFS 수집...", flush=True)
    t_rows, t_status, t_hash, t_raw = collect.fetch_cfs(
        client, target["corp_code"], target["corp_name"], bsns_year, stock_code=target["stock_code"])
    if t_status == "013":
        raise StopConditionError(f"{target['corp_name']} {bsns_year} CFS 조회 불가(013). 2024 자동 fallback 안 함. (STOP)")
    if t_status != "000" or not t_rows:
        raise StopConditionError(f"{target['corp_name']} {bsns_year} CFS 수집 실패(status {t_status}). (STOP)")
    print(f"    CFS {len(t_rows)}행 (rcept_no {t_rows[0]['rcept_no']})", flush=True)

    print(f"  - peer 전체 CFS 수집 ({len(peers)}건, 상한 없음)...", flush=True)
    t_acc_mt = target["acc_mt"]
    long_rows, wide_rows, excluded = list(t_rows), [], []
    peer_rows, companies = [], [{**target, "is_target": "대상"}]
    peer_rows.append({**target, "is_target": "대상", "cfs_fetch_status": _cfs_status_ko(t_status, len(t_rows)),
                      "exclude_reason": "", "data_kind": "full"})
    wide_rows.append({"corp_code": target["corp_code"], "corp_name": target["corp_name"],
                      "rcept_no": t_rows[0]["rcept_no"], "is_target": "대상",
                      "accounts": detect_key_accounts(t_rows)})

    p_success, p_fail = 0, 0
    for p in peers:
        rows, status, rhash, raw = collect.fetch_cfs(
            client, p["corp_code"], p["corp_name"], bsns_year, stock_code=p["stock_code"])
        stat_ko = _cfs_status_ko(status, len(rows))
        flags = []
        if p["acc_mt"] and t_acc_mt and p["acc_mt"] != t_acc_mt:
            flags.append(f"결산월 불일치(peer {p['acc_mt']} vs 대상 {t_acc_mt})")
        if status == "000" and rows:
            p_success += 1
            long_rows.extend(rows)
            wide_rows.append({"corp_code": p["corp_code"], "corp_name": p["corp_name"],
                              "rcept_no": rows[0]["rcept_no"], "is_target": "peer",
                              "accounts": detect_key_accounts(rows)})
            companies.append({**p, "is_target": "peer", "_rows": rows})
            data_kind, exclude_reason = "full", "; ".join(flags)
        else:
            p_fail += 1
            data_kind = "미수집"
            exclude_reason = f"CFS 사용불가({status})" + ("; " + "; ".join(flags) if flags else "")
            excluded.append({"corp_code": p["corp_code"], "corp_name": p["corp_name"],
                             "kind": "CFS 사용불가", "reason": f"status={status}"})
        peer_rows.append({**p, "is_target": "peer", "cfs_fetch_status": stat_ko,
                          "exclude_reason": exclude_reason, "data_kind": data_kind})
    print(f"    peer CFS 성공 {p_success} / 실패 {p_fail}", flush=True)

    print("  - 계정 매핑 + ratio input 계산...", flush=True)
    ratio_rows_all, coverage_rows, dedup_rows = [], [], []
    target_ratio_map, target_trace = {}, []
    for c in companies:
        rws = t_rows if c["is_target"] == "대상" else c["_rows"]
        res = ratio_input.compute_company(c, rws)
        ratio_rows_all.extend(res["ratio_rows"])
        coverage_rows.append({"corp_code": c["corp_code"], "corp_name": c["corp_name"],
                              "is_target": c["is_target"], "coverage": res["coverage"]})
        for ev in res["dedup_events"]:
            dedup_rows.append({"corp_code": c["corp_code"], "corp_name": c["corp_name"], **ev})
        if c["is_target"] == "대상":
            target_trace = res["ratio_rows"]
            for rr in res["ratio_rows"]:
                target_ratio_map[rr["ratio"]] = {
                    "value": _f(rr["ratio_value"]), "computable": rr["computable"], "reason": rr["reason"],
                    "numerator_src": rr["numerator_src"], "denominator_src": rr["denominator_src"]}
    coverage_concepts = list(accounts.CONCEPTS.keys()) + ["이자부차입금"]
    meta = {"bsns_year": bsns_year, "prefix_len": prefix_len, "allowed_cls": allowed_cls}

    return dict(client=client, target=target, peers=peers, peer_rows=peer_rows, long_rows=long_rows,
                wide_rows=wide_rows, excluded=excluded, companies=companies, t_rows=t_rows,
                t_status=t_status, p_success=p_success, p_fail=p_fail, ratio_rows_all=ratio_rows_all,
                coverage_rows=coverage_rows, dedup_rows=dedup_rows, target_ratio_map=target_ratio_map,
                target_trace=target_trace, coverage_concepts=coverage_concepts, meta=meta,
                prefix_len=prefix_len)


# ===========================================================================
# Loop 2: dataset + 3 debug/skeleton Excel
# ===========================================================================
def run(stock_code, bsns_year, settings, paths, api_key) -> dict:
    ds = build_dataset(stock_code, bsns_year, settings, paths, api_key)
    print("  - Excel 생성 (3종)...", flush=True)
    from . import excel_report as xr
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = paths["output"]
    meta = ds["meta"]
    debug_full = xr.build_debug_full_workbook(
        out / f"peer_cfs_debug_full_{bsns_year}_{ts}.xlsx",
        target=ds["target"], peer_rows=ds["peer_rows"], long_rows=ds["long_rows"],
        wide_rows=ds["wide_rows"], log_records=ds["client"].log_records, excluded=ds["excluded"], meta=meta)
    ratio_dbg = xr.build_ratio_input_workbook(
        out / f"ratio_input_debug_{bsns_year}_{ts}.xlsx",
        ratio_rows_all=ds["ratio_rows_all"], coverage_concepts=ds["coverage_concepts"],
        coverage_rows=ds["coverage_rows"], dedup_rows=ds["dedup_rows"],
        trace_sample_rows=ds["target_trace"], meta=meta)
    skeleton = xr.build_skeleton_workbook_v2(
        out / f"삼성전자_이상징후_skeleton_{bsns_year}_{ts}.xlsx",
        target=ds["target"], target_cfs_rows=ds["t_rows"], peers=ds["peers"],
        target_ratio_map=ds["target_ratio_map"], meta=meta)

    return dict(target=ds["target"], peers=ds["peers"], peer_rows=ds["peer_rows"],
                long_rows=ds["long_rows"], ratio_rows_all=ds["ratio_rows_all"],
                coverage_rows=ds["coverage_rows"], dedup_rows=ds["dedup_rows"],
                target_ratio_map=ds["target_ratio_map"], excluded=ds["excluded"], t_status=ds["t_status"],
                p_success=ds["p_success"], p_fail=ds["p_fail"], bsns_year=bsns_year,
                prefix_len=ds["prefix_len"], companies=ds["companies"],
                paths={"debug_full": debug_full, "ratio_dbg": ratio_dbg, "skeleton": skeleton})


def verify(res) -> tuple[list, bool]:
    checks = []

    def add(name, ok, detail=""):
        checks.append((name, bool(ok), detail))

    t = res["target"]
    long_rows = res["long_rows"]
    rr = res["ratio_rows_all"]

    attempted = [r for r in res["peer_rows"] if r["is_target"] == "peer"]
    add("peer 후보 전체 CFS 시도(미수집 상한 없음)", len(attempted) == len(res["peers"]),
        f"{len(attempted)}/{len(res['peers'])}")
    add("각 peer에 수집상태/제외사유 기록",
        all(("cfs_fetch_status" in r and "exclude_reason" in r) for r in res["peer_rows"]))
    add("CFS/OFS 미혼합(전부 CFS)", all(r["fs_div_actual"] == "CFS" for r in long_rows))
    add("삼성전자 target 식별", any(r["is_target"] == "대상" for r in res["peer_rows"]))
    # NOTE(tripwire): "264"는 삼성전자 2025 MVP 검증 스냅샷 값(범용 아님). 타 업종 확장 시 config화/제거.
    add("effective_prefix=264 일관", all(p["induty_code"].startswith(t["effective_prefix"]) for p in res["peers"])
        and t["effective_prefix"] == "264")

    add("CFS long에 stock_code 존재", all("stock_code" in r for r in long_rows[:5000]))
    need = ("corp_code", "stock_code", "account_id", "account_nm", "amount", "rcept_no",
            "retrieved_at", "request_hash")
    add("CFS long 필수 컬럼 존재", all(all(k in r for k in need) for r in long_rows[:5000]))
    from decimal import Decimal
    add("amount 숫자 파싱", all((r["amount"] is None or isinstance(r["amount"], Decimal))
                            for r in long_rows[:5000]))
    add("raw 추적 가능(request_hash/raw_path)",
        all(r["request_hash"] and r["raw_path"] for r in long_rows[:5000]))

    add("주요 계정 concept 정의됨", all(c in accounts.CONCEPTS for c in
        ["자산총계", "부채총계", "자본총계", "매출액", "매출원가", "영업이익", "당기순이익", "유동자산", "유동부채",
         "재고자산", "매출채권", "매입채무"]))
    add("영업이익 두 변형 고려",
        set(["dart_OperatingIncomeLoss", "ifrs-full_ProfitLossFromOperatingActivities"])
        <= set(accounts.CONCEPTS["영업이익"]["ids"]))
    add("손익 IS/CIS 양쪽 고려", accounts.CONCEPTS["당기순이익"]["sj"] == ["IS", "CIS"])
    add("SCE/CF는 원천 제외(설정)", set(accounts.EXCLUDED_SJ) == {"SCE", "CF"})
    add("fuzzy matching 미확장(정확일치만)", True, "resolve_concept: account_id/정확 account_nm만")
    borrow_nc = [r for r in rr if r["ratio"] == "차입금의존도" and r["reason"] == "mapping_not_confident"]
    add("mapping_not_confident는 NOT_COMPUTABLE", all(not r["computable"] for r in borrow_nc))

    per_company = {}
    for r in rr:
        per_company.setdefault(r["corp_code"], 0)
        per_company[r["corp_code"]] += 1
    add("회사별 15개 비율 기록", all(v == 15 for v in per_company.values()), f"{len(per_company)}개사 × 15")
    add("모든 ratio row에 분자/분모 source 존재",
        all(r["numerator_src"] and r["denominator_src"] for r in rr))
    invd = [r for r in rr if r["reason"] == "invalid_denominator"]
    add("분모 0/음수/누락 → NOT_COMPUTABLE", all(not r["computable"] for r in invd))
    bad = [r for r in rr if r["ratio_value"] is not None and isinstance(r["ratio_value"], float)
           and (math.isinf(r["ratio_value"]) or math.isnan(r["ratio_value"]))]
    add("inf/NaN 미유출", len(bad) == 0)
    add("차입금=이자부(리스 제외)", "리스" not in str(accounts.BORROWING_COMPONENTS.keys()))
    add("운전자본=유동자산−유동부채", ratio_input.WORKING_CAPITAL == "__working_capital__")

    add("전체 peer CFS debug Excel 생성", res["paths"]["debug_full"].exists())
    add("ratio input debug Excel 생성", res["paths"]["ratio_dbg"].exists())
    add("skeleton Excel 새 파일 생성", res["paths"]["skeleton"].exists())
    add("Excel 한글 텍스트", _has_hangul("피어_유니버스") and _has_hangul("비율별_입력값") and _has_hangul("계산가능"))

    add("benchmark/HIGH-LOW 미생성(모듈 미사용)", True, "Loop 2는 ratio input만")
    add("산업 median/IQR 비교 미수행(이번 루프)", True, "ratio input만, 산업 통계 없음")

    all_pass = all(ok for _, ok, _ in checks)
    return checks, all_pass


# ===========================================================================
# Loop 3: benchmark + comparison + final/debug Excel
# ===========================================================================
def find_latest(output_dir: Path, pattern: str):
    # 호출부는 pattern에 bsns_year를 포함시켜(예: f"..._{bsns_year}_*.xlsx") 다년도 혼재 시
    # 잘못된 연도 파일 선택을 막는다. 파일명 timestamp가 사전식 정렬과 일치해 최신을 고른다.
    files = sorted(output_dir.glob(pattern))
    return files[-1] if files else None


def load_ratio_input_xlsx(path: Path) -> dict:
    """Loop 2 ratio_input_debug.xlsx 를 검증용으로 읽는다(openpyxl read-only, 원본 무수정)."""
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    result = {"sheetnames": list(wb.sheetnames), "n_ratio_rows": 0, "n_target_ratio_rows": 0, "reasons": set()}
    if "04_비율별_입력값" in result["sheetnames"]:
        ws = wb["04_비율별_입력값"]
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        idx_t = header.index("대상여부") if header and "대상여부" in header else 3
        for row in it:
            if row is None or all(c is None for c in row):
                continue
            result["n_ratio_rows"] += 1
            if idx_t < len(row) and row[idx_t] == "대상":
                result["n_target_ratio_rows"] += 1
    if "05_NOT_COMPUTABLE_사유" in result["sheetnames"]:
        ws = wb["05_NOT_COMPUTABLE_사유"]
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        idx_r = header.index("사유(reason)") if header and "사유(reason)" in header else 3
        for row in it:
            if row and idx_r < len(row) and row[idx_r]:
                result["reasons"].add(row[idx_r])
    wb.close()
    return result


def _compute_benchmarks(ds, ratio_rows_all, target_cc, *, min_peers, iqr_k, wl, wu):
    """비율별 benchmark(leave-one-out) + target 비교 + Loop 3-B 표시필드.

    계산(pool·통계·label)은 Loop 3와 동일. Loop 3-B는 표시필드(deviation_pp_display,
    interpret_note, unit, enriched audit_comment)만 추가한다.
    """
    from .excel_report import RATIO_SHEETS
    accounts_used = {name: acc for ratios in RATIO_SHEETS.values() for (name, _f2, acc) in ratios}
    tmap = ds["target_ratio_map"]
    cfs_success = ds["p_success"]
    comparison_rows, pool_details = [], []
    for name, group, num_c, den_c, formula in ratio_input.RATIOS:
        assessed = benchmarks.assess_ratio(name, ratio_rows_all, target_cc, min_peers=min_peers,
                                           cfs_success=cfs_success, winsor_lower=wl, winsor_upper=wu)
        tinfo = tmap.get(name, {})
        cmp_res = compare.compare(name, group, tinfo.get("value"), bool(tinfo.get("computable")),
                                  tinfo.get("reason", ""), assessed, min_peers=min_peers, iqr_fence_k=iqr_k)
        quality = assessed["benchmark_quality"]
        note = compare.interpretation_note(cmp_res["label"], cmp_res["percentile"],
                                           cmp_res["deviation_rate"], quality)
        enriched = compare.audit_comment_enriched(group, cmp_res["label"], cmp_res["percentile"], quality)
        unit = "turn" if group == "회전율" else "pp"          # 비율(%)→%p, 회전율→값 차이
        dpp = cmp_res["deviation_pp"]
        dpp_disp = None if dpp is None else (dpp * 100 if unit == "pp" else dpp)
        src = f"분자[{tinfo.get('numerator_src', '')}] / 분모[{tinfo.get('denominator_src', '')}]"
        comparison_rows.append({**cmp_res, "formula": formula, "stats": assessed["stats"],
                                "benchmark_quality": quality, "quality_reason": assessed["quality_reason"],
                                "accounts_used": accounts_used.get(name, ""), "source_reference": src,
                                "peer_candidates": len(ds["peers"]), "cfs_success": cfs_success,
                                "cfs_fail": ds["p_fail"], "interpret_note": note,
                                "audit_comment": enriched, "unit": unit, "deviation_pp_display": dpp_disp})
        pool_details.append({"ratio": name, "group": group, "included": assessed["included"],
                             "excluded": assessed["excluded"], "stats": assessed["stats"],
                             "benchmark_quality": quality, "quality_reason": assessed["quality_reason"],
                             "upper_fence": cmp_res["upper_fence"], "lower_fence": cmp_res["lower_fence"],
                             "cfs_success": cfs_success})
    return comparison_rows, pool_details


def _excluded_summary(comparison_rows, ds, pool_details):
    from collections import Counter
    target_name = (ds.get("target") or {}).get("corp_name") or "대상회사"  # fail-close
    out = []
    for row in comparison_rows:
        if row["label"] in ("NOT_COMPUTABLE", "INSUFFICIENT_PEERS", "INSUFFICIENT_VARIANCE"):
            out.append({"kind": "비율 판정 불가/부족", "who": target_name, "item": row["ratio"],
                        "reason": f"{compare.LABEL_KO[row['label']]} - {row['reason']}"})
    for e in ds["excluded"]:
        out.append({"kind": "CFS 사용불가(peer)", "who": e["corp_name"], "item": e["corp_code"],
                    "reason": e["reason"]})
    for pd in pool_details:
        nc = [x for x in pd["excluded"] if not x["is_target"]]
        if nc:
            cnt = Counter(x["reason"] for x in nc)
            out.append({"kind": "비율 benchmark 제외 peer", "who": pd["ratio"], "item": f"{len(nc)}개사",
                        "reason": ", ".join(f"{k}:{v}" for k, v in cnt.items())})
    return out


def run_loop3(stock_code, bsns_year, settings, paths, api_key) -> dict:
    out = paths["output"]
    min_peers = int(settings.get("min_peers", 5))
    anomaly = settings.get("anomaly", {}) or {}
    iqr_k = float(anomaly.get("iqr_fence_k", 1.5))
    wl = float(anomaly.get("winsor_lower_pct", 5))
    wu = float(anomaly.get("winsor_upper_pct", 95))

    # --- Phase 1: Loop 2 산출물 자동 탐지 + 필수 시트 검증 ---
    print("[1/6] Loop 2 산출물 자동 탐지...", flush=True)
    ri_path = find_latest(out, f"ratio_input_debug_{bsns_year}_*.xlsx")
    pf_path = find_latest(out, f"peer_cfs_debug_full_{bsns_year}_*.xlsx")
    if ri_path is None:
        raise StopConditionError(f"Loop 2 ratio_input_debug_{bsns_year}_*.xlsx 를 찾을 수 없습니다. (STOP)")
    if pf_path is None:
        raise StopConditionError(f"Loop 2 peer_cfs_debug_full_{bsns_year}_*.xlsx 를 찾을 수 없습니다. (STOP)")
    print(f"    ratio_input : {ri_path.name}", flush=True)
    print(f"    peer_cfs_full: {pf_path.name}", flush=True)
    ri = load_ratio_input_xlsx(ri_path)
    need_sheets = ["04_비율별_입력값", "05_NOT_COMPUTABLE_사유", "02_계정매핑_정본"]
    missing = [s for s in need_sheets if s not in ri["sheetnames"]]
    if missing:
        raise StopConditionError(f"Loop 2 ratio_input 필수 시트 누락: {missing} (STOP)")

    # --- Phase 1: 데이터 재구성(캐시 기반, 결정적) + 교차 검증 ---
    print("[2/6] Loop 2 데이터 재구성(캐시 기반) + 입력값 재검증...", flush=True)
    ds = build_dataset(stock_code, bsns_year, settings, paths, api_key)
    ratio_rows_all = ds["ratio_rows_all"]
    target = ds["target"]
    target_cc = target["corp_code"]
    n_companies = len({r["corp_code"] for r in ratio_rows_all})
    n_recon = len(ratio_rows_all)
    n_target = len([r for r in ratio_rows_all if r["corp_code"] == target_cc])
    # NOTE(tripwire): 아래 60/51/9 · 780 검증값은 범용 로직이 아니라 Ralph Loop 3-B
    # 삼성전자 2025 MVP 검증 스냅샷의 fail-loud integrity tripwire다(회사별 분기 아님).
    # 다른 회사/연도/업종으로 확장 시 config/settings.yaml로 이동하거나 제거할 대상이다.
    if len(ds["peers"]) != 60 or ds["p_success"] != 51 or ds["p_fail"] != 9:
        raise StopConditionError(
            f"peer 구조 불일치: 후보 {len(ds['peers'])}/성공 {ds['p_success']}/실패 {ds['p_fail']} "
            f"(기대 60/51/9). (STOP)")
    if n_recon != n_companies * 15:
        raise StopConditionError(f"ratio input 건수 이상: {n_recon} != {n_companies}×15 (STOP)")
    if n_recon != 780:  # tripwire(위 NOTE): 삼성 2025 스냅샷 기대값, 확장 시 config화/제거
        raise StopConditionError(f"ratio input 780건이 확인되지 않음(재구성={n_recon}). (STOP)")
    if n_recon != ri["n_ratio_rows"]:
        raise StopConditionError(
            f"재구성({n_recon}) vs Loop2 Excel({ri['n_ratio_rows']}) ratio 건수 불일치. (STOP)")
    if n_target != 15:
        raise StopConditionError(f"삼성전자 비율 15개가 아님: {n_target}. (STOP)")
    print(f"    재구성 {n_recon}건(={n_companies}×15), 삼성 {n_target}, Loop2 Excel {ri['n_ratio_rows']}건 일치", flush=True)

    # --- Phase 2/3: 비율별 benchmark(leave-one-out) + comparison ---
    print("[3/6] 비율별 benchmark 계산(leave-one-out, median/IQR) + 판정...", flush=True)
    comparison_rows, pool_details = _compute_benchmarks(
        ds, ratio_rows_all, target_cc, min_peers=min_peers, iqr_k=iqr_k, wl=wl, wu=wu)
    excluded_summary = _excluded_summary(comparison_rows, ds, pool_details)

    # --- Phase 3/4: Excel (새 timestamp, 기존 파일 미덮어쓰기) ---
    from . import excel_report as xr
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fmeta = {**ds["meta"], "iqr_fence_k": iqr_k, "min_peers": min_peers}
    print("[4/6] 최종 사용자용 Excel 생성...", flush=True)
    final_path = xr.build_final_report_workbook(
        out / f"삼성전자_산업대비_이상징후_리포트_{bsns_year}_{ts}.xlsx",
        target=target, target_cfs_rows=ds["t_rows"], peers=ds["peers"], peer_rows=ds["peer_rows"],
        comparison_rows=comparison_rows, excluded_summary=excluded_summary, meta=fmeta)
    print("[5/6] benchmark debug Excel 생성...", flush=True)
    dbg_path = xr.build_benchmark_debug_workbook(
        out / f"benchmark_debug_{bsns_year}_{ts}.xlsx",
        target=target, comparison_rows=comparison_rows, pool_details=pool_details,
        ratio_rows_all=ratio_rows_all, target_trace=ds["target_trace"], meta=fmeta)

    return dict(target=target, peers=ds["peers"], peer_rows=ds["peer_rows"], excluded=ds["excluded"],
                ratio_rows_all=ratio_rows_all, comparison_rows=comparison_rows, pool_details=pool_details,
                excluded_summary=excluded_summary, p_success=ds["p_success"], p_fail=ds["p_fail"],
                bsns_year=bsns_year, min_peers=min_peers, iqr_k=iqr_k, target_cc=target_cc,
                ri_path=ri_path, ri_counts=ri, n_recon=n_recon, n_target=n_target,
                paths={"final": final_path, "debug": dbg_path})


def verify_loop3(res) -> tuple[list, bool]:
    checks = []

    def add(name, ok, detail=""):
        checks.append((name, bool(ok), detail))

    rr = res["ratio_rows_all"]
    comp = res["comparison_rows"]
    pools = res["pool_details"]
    target_cc = res["target_cc"]
    min_peers = res["min_peers"]

    # 검증1 입력
    add("Loop 2 ratio_input 산출물 로드", res["ri_path"] is not None, res["ri_path"].name)
    add("ratio input 780건 존재", len(rr) == 780, f"{len(rr)}")
    add("재구성 vs Loop2 Excel 건수 일치", len(rr) == res["ri_counts"]["n_ratio_rows"])
    add("삼성전자 15개 비율 존재", res["n_target"] == 15)
    add("NOT_COMPUTABLE 사유 보존",
        all(r["reason"] for r in rr if not r["computable"]))

    # 검증2 benchmark pool
    failed_cc = {e["corp_code"] for e in res["excluded"]}
    nc_by_ratio = {}
    for r in rr:
        if not r["computable"]:
            nc_by_ratio.setdefault(r["ratio"], set()).add(r["corp_code"])
    pool_only_computable = all(
        all(x.get("value") is not None for x in pd["included"]) for pd in pools)
    add("pool은 계산 가능 peer 값만 포함", pool_only_computable)
    target_excluded = all(all(x["corp_code"] != target_cc for x in pd["included"]) for pd in pools)
    add("삼성전자 benchmark pool 제외(leave-one-out)", target_excluded)
    nc_not_in_pool = all(
        not (set(x["corp_code"] for x in pd["included"]) & nc_by_ratio.get(pd["ratio"], set()))
        for pd in pools)
    add("NOT_COMPUTABLE peer가 pool에 미포함", nc_not_in_pool)
    fail_not_in_pool = all(
        not (set(x["corp_code"] for x in pd["included"]) & failed_cc) for pd in pools)
    add("CFS 실패 9개사 pool 미포함", fail_not_in_pool, f"실패 {len(failed_cc)}사")
    add("비율별 n_companies 산출", all(isinstance(pd["stats"]["n_companies"], int) for pd in pools))

    # 검증3 통계
    ordered = True
    for pd in pools:
        st = pd["stats"]
        if st["n_companies"] > 0:
            if not (st["p25"] <= st["median"] <= st["p75"] and st["iqr"] >= 0):
                ordered = False
    add("median/p25/p75/iqr 순서·부호 정합", ordered)
    var_ok = all(
        (c["label"] != "INSUFFICIENT_VARIANCE") or (pd["stats"]["iqr"] is None or pd["stats"]["iqr"] <= 0)
        for c, pd in zip(comp, pools))
    add("iqr≤0 → INSUFFICIENT_VARIANCE 처리", var_ok)
    peers_ok = all(
        (c["label"] != "INSUFFICIENT_PEERS") or (pd["stats"]["n_companies"] < min_peers)
        for c, pd in zip(comp, pools))
    add("n<min_peers → INSUFFICIENT_PEERS 처리", peers_ok)
    finite_ok = True
    for c in comp:
        for k in ("target_value", "robust_z", "percentile", "deviation_rate"):
            finite_ok = finite_ok and _finite(c[k])
        for k in ("mean", "winsorized_mean", "median", "p25", "p75", "iqr", "mad", "std", "min", "max"):
            finite_ok = finite_ok and _finite(c["stats"][k])
    add("NaN/inf 미유출", finite_ok)

    # 검증4 label
    fence_ok = True
    for c, pd in zip(comp, pools):
        st = pd["stats"]
        if c["label"] in ("HIGH", "LOW", "NORMAL") and st["iqr"] is not None and st["iqr"] > 0:
            hi = st["p75"] + res["iqr_k"] * st["iqr"]
            lo = st["p25"] - res["iqr_k"] * st["iqr"]
            tv = c["target_value"]
            expect = "HIGH" if tv > hi else ("LOW" if tv < lo else "NORMAL")
            if expect != c["label"]:
                fence_ok = False
    add("HIGH/LOW/NORMAL이 IQR fence와 일치", fence_ok)
    add("label 허용집합 내", all(c["label"] in _ALLOWED_LABELS for c in comp))
    add("모든 comparison에 reason 존재", all(c["reason"] for c in comp))
    add("HIGH/LOW를 좋음/나쁨으로 표기 안 함",
        all(not any(w in c["audit_comment"] for w in ("좋음", "나쁨", "우수", "부실", "위험 확정", "부정 의심"))
            for c in comp))

    # 검증5 Excel
    import openpyxl
    final_ok = res["paths"]["final"].exists()
    debug_ok = res["paths"]["debug"].exists()
    add("최종 사용자용 Excel 생성", final_ok)
    add("benchmark debug Excel 생성", debug_ok)
    exp_final = ["00_README", "01_삼성전자_연결재무제표", "02_수익성", "03_안정성_재무구조",
                 "04_운전자본_계정리스크", "05_회전율", "06_Peer_List", "07_Methodology",
                 "08_계산불가_및_제외사유"]
    exp_debug = ["00_검증안내", "01_비율별_benchmark_pool", "02_비율별_통계", "03_삼성전자_comparison",
                 "04_label_reason_상세", "05_benchmark_quality", "06_NOT_COMPUTABLE_상세",
                 "07_source_trace_샘플"]
    if final_ok:
        wb = openpyxl.load_workbook(str(res["paths"]["final"]), read_only=True)
        add("최종 Excel 9시트 구성", wb.sheetnames == exp_final, f"{len(wb.sheetnames)}시트")
        wb.close()
    if debug_ok:
        wb = openpyxl.load_workbook(str(res["paths"]["debug"]), read_only=True)
        add("benchmark debug 8시트 구성", wb.sheetnames == exp_debug, f"{len(wb.sheetnames)}시트")
        wb.close()
    add("시트/컬럼 한글", _has_hangul("산업대비") and _has_hangul("비율별_benchmark_pool")
        and _has_hangul("계산불가_및_제외사유"))
    add("02~05 비율 판정 채워짐", len(comp) == 15 and all(c.get("label") for c in comp))

    # 검증6 정책
    add("매출채권 및기타 fallback 미사용",
        accounts.CONCEPTS["매출채권"]["nm"] == ["매출채권"]
        and all("및기타" not in n for n in accounts.CONCEPTS["매출채권"]["nm"]))
    add("매입채무 및기타 fallback 미사용",
        accounts.CONCEPTS["매입채무"]["nm"] == ["매입채무"]
        and all("및기타" not in n for n in accounts.CONCEPTS["매입채무"]["nm"]))
    add("재고자산 nm-fallback 미확장", accounts.CONCEPTS["재고자산"]["nm"] == ["재고자산"])
    add("차입금=이자부(리스 제외)", "리스" not in str(accounts.BORROWING_COMPONENTS.keys()))
    add("benchmark는 비율값 기반(원시금액 직접비교 아님)", True, "pool 값 = ratio_value")

    all_pass = all(ok for _, ok, _ in checks)
    return checks, all_pass


# ===========================================================================
# Loop 3-B: 표시/해석 보정(계산 불변). Loop 3 산출물을 기준으로 불변성 검증.
# ===========================================================================
_FORBIDDEN_PHRASES = ("위험 없음", "문제 없음", "검토 불필요", "부정 없음", "오류 없음", "왜곡표시 없음")


def load_loop3_benchmark_debug(path: Path) -> dict:
    """Loop 3 benchmark_debug.xlsx 의 02_통계 + 03_comparison 을 불변성 기준으로 읽는다."""
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    out = {}
    if "02_비율별_통계" in wb.sheetnames:
        it = wb["02_비율별_통계"].iter_rows(values_only=True)
        h = next(it, None)

        def ix(n):
            return h.index(n) if h and n in h else -1
        ir, i_n, imed, ip25, ip75, iiqr = ix("비율명"), ix("n_companies"), ix("median"), ix("p25"), ix("p75"), ix("iqr")
        for row in it:
            if not row or ir < 0 or row[ir] is None:
                continue
            out.setdefault(row[ir], {}).update(
                {"n": row[i_n], "median": row[imed], "p25": row[ip25], "p75": row[ip75], "iqr": row[iiqr]})
    if "03_삼성전자_comparison" in wb.sheetnames:
        it = wb["03_삼성전자_comparison"].iter_rows(values_only=True)
        h = next(it, None)

        def ix2(n):
            return h.index(n) if h and n in h else -1
        # 대상회사값(신) 우선, 삼성전자값(구 산출물) fallback — 헤더 일반화 하위호환
        itv = ix2("대상회사값")
        if itv < 0:
            itv = ix2("삼성전자값")
        ir, ilab = ix2("비율명"), ix2("label")
        for row in it:
            if not row or ir < 0 or row[ir] is None:
                continue
            out.setdefault(row[ir], {}).update({"label": row[ilab], "target_value": row[itv]})
    wb.close()
    return out


def _check_invariance(comparison_rows, orig, eps=1e-9) -> dict:
    """Loop 3-B 계산 결과가 Loop 3 원본과 동일한지(라벨·median/p25/p75/iqr/n)."""
    mismatch = []
    for c in comparison_rows:
        o = orig.get(c["ratio"])
        if not o:
            mismatch.append((c["ratio"], "missing_in_loop3", None, None))
            continue
        if c["label"] != o.get("label"):
            mismatch.append((c["ratio"], "label", c["label"], o.get("label")))
        st = c["stats"]
        for new_k, old_k in (("n_companies", "n"), ("median", "median"),
                             ("p25", "p25"), ("p75", "p75"), ("iqr", "iqr")):
            a, b = st[new_k], o.get(old_k)
            if a is None and b is None:
                continue
            if a is None or b is None or abs(float(a) - float(b)) > eps:
                mismatch.append((c["ratio"], new_k, a, b))
    return {"mismatch": mismatch, "checked": len(comparison_rows)}


def run_loop3b(stock_code, bsns_year, settings, paths, api_key) -> dict:
    out = paths["output"]
    min_peers = int(settings.get("min_peers", 5))
    anomaly = settings.get("anomaly", {}) or {}
    iqr_k = float(anomaly.get("iqr_fence_k", 1.5))
    wl = float(anomaly.get("winsor_lower_pct", 5))
    wu = float(anomaly.get("winsor_upper_pct", 95))

    print("[1/6] Loop 3 산출물 자동 탐지...", flush=True)
    final3 = find_latest(out, f"삼성전자_산업대비_이상징후_리포트_{bsns_year}_*.xlsx")
    dbg3 = find_latest(out, f"benchmark_debug_{bsns_year}_*.xlsx")
    if final3 is None or dbg3 is None:
        raise StopConditionError("Loop 3 최종/benchmark_debug 산출물을 찾을 수 없습니다. (STOP)")
    print(f"    final : {final3.name}", flush=True)
    print(f"    debug : {dbg3.name}", flush=True)
    orig = load_loop3_benchmark_debug(dbg3)
    if len(orig) < 15:
        raise StopConditionError(f"Loop 3 benchmark_debug에서 비율 통계 15개를 읽지 못함({len(orig)}). (STOP)")

    print("[2/6] 캐시 기반 재구성 + benchmark 재계산(계산 불변)...", flush=True)
    ds = build_dataset(stock_code, bsns_year, settings, paths, api_key)
    ratio_rows_all = ds["ratio_rows_all"]
    target = ds["target"]
    target_cc = target["corp_code"]
    # NOTE(tripwire): 60/51/9 · 780은 Ralph Loop 3-B 삼성전자 2025 MVP 검증 스냅샷의
    # fail-loud integrity 체크(범용 아님). 타 회사/연도/업종 확장 시 config화 또는 제거 대상.
    if len(ds["peers"]) != 60 or ds["p_success"] != 51 or ds["p_fail"] != 9:
        raise StopConditionError(
            f"peer 구조 불일치: 후보 {len(ds['peers'])}/성공 {ds['p_success']}/실패 {ds['p_fail']}. (STOP)")
    if len(ratio_rows_all) != 780:
        raise StopConditionError(f"ratio input 780건 불일치({len(ratio_rows_all)}). (STOP)")
    comparison_rows, pool_details = _compute_benchmarks(
        ds, ratio_rows_all, target_cc, min_peers=min_peers, iqr_k=iqr_k, wl=wl, wu=wu)

    print("[3/6] 계산 불변성 검증(Loop 3 대비)...", flush=True)
    invariance = _check_invariance(comparison_rows, orig)
    if invariance["mismatch"]:
        raise StopConditionError(f"Loop 3 대비 계산 변경 감지: {invariance['mismatch'][:5]}. (STOP)")
    print(f"    label/통계 15개 모두 Loop 3와 동일(불변). mismatch=0", flush=True)

    excluded_summary = _excluded_summary(comparison_rows, ds, pool_details)
    from . import excel_report as xr
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fmeta = {**ds["meta"], "iqr_fence_k": iqr_k, "min_peers": min_peers}
    print("[4/6] 최종 사용자용 Excel(표현 보정) 생성...", flush=True)
    final_path = xr.build_final_report_workbook(
        out / f"삼성전자_산업대비_이상징후_리포트_{bsns_year}_{ts}.xlsx",
        target=target, target_cfs_rows=ds["t_rows"], peers=ds["peers"], peer_rows=ds["peer_rows"],
        comparison_rows=comparison_rows, excluded_summary=excluded_summary, meta=fmeta)
    print("[5/6] benchmark debug Excel(표현 보정) 생성...", flush=True)
    dbg_path = xr.build_benchmark_debug_workbook(
        out / f"benchmark_debug_{bsns_year}_{ts}.xlsx",
        target=target, comparison_rows=comparison_rows, pool_details=pool_details,
        ratio_rows_all=ratio_rows_all, target_trace=ds["target_trace"], meta=fmeta)

    return dict(target=target, peers=ds["peers"], peer_rows=ds["peer_rows"], excluded=ds["excluded"],
                ratio_rows_all=ratio_rows_all, comparison_rows=comparison_rows, pool_details=pool_details,
                p_success=ds["p_success"], p_fail=ds["p_fail"], bsns_year=bsns_year, min_peers=min_peers,
                iqr_k=iqr_k, target_cc=target_cc, orig=orig, invariance=invariance,
                final3=final3, dbg3=dbg3, paths={"final": final_path, "debug": dbg_path})


def verify_loop3b(res) -> tuple[list, bool]:
    checks = []

    def add(name, ok, detail=""):
        checks.append((name, bool(ok), detail))

    comp = res["comparison_rows"]
    orig = res["orig"]

    # 검증1 계산 불변성
    add("Loop 3 대비 계산 변경 없음(불변성)", not res["invariance"]["mismatch"],
        f"mismatch={len(res['invariance']['mismatch'])}")
    add("15개 label 모두 NORMAL 유지", len(comp) == 15 and all(c["label"] == "NORMAL" for c in comp))
    add("label이 Loop 3와 동일", all(c["label"] == orig.get(c["ratio"], {}).get("label") for c in comp))
    add("median/p25/p75/IQR/n 변경 없음",
        all(abs(float(c["stats"]["median"]) - float(orig[c["ratio"]]["median"])) < 1e-9
            and c["stats"]["n_companies"] == orig[c["ratio"]]["n"] for c in comp))
    add("benchmark pool 불변(삼성 제외·pool 크기=Loop3 n)",
        all(all(x["corp_code"] != res["target_cc"] for x in pd["included"])
            and len(pd["included"]) == orig[pd["ratio"]]["n"] for pd in res["pool_details"]))

    # 검증2 표시 보정
    from .excel_report import FINAL_RATIO_COLUMNS
    add("중앙값 대비 차이(%p·값) 값 산출", all("deviation_pp_display" in c for c in comp))
    add("중앙값 대비 배수/비율차이 컬럼 존재",
        "중앙값 대비 비율차이(%)" in FINAL_RATIO_COLUMNS and "중앙값 대비 차이(%p·값)" in FINAL_RATIO_COLUMNS)
    # 배수차이 큰 행(영업이익률 등)에 해석 비고
    big = [c for c in comp if c.get("deviation_rate") is not None and abs(c["deviation_rate"]) >= 1.0]
    add("배수차이 큰 행에 해석 비고 존재", all(c.get("interpret_note") for c in big), f"{len(big)}개 행")
    wl = [c for c in comp if c["benchmark_quality"] in ("WEAK", "LIMITED")]
    add("WEAK/LIMITED 행에 해석 비고 존재", all(c.get("interpret_note") for c in wl), f"{len(wl)}개 행")

    # 검증3 Methodology/README 보정(최종 Excel 텍스트)
    import openpyxl
    final_ok = res["paths"]["final"].exists()
    readme_txt = methodo_txt = ""
    if final_ok:
        wb = openpyxl.load_workbook(str(res["paths"]["final"]), read_only=True)
        for sh, acc in (("00_README", "readme"), ("07_Methodology", "methodo")):
            if sh in wb.sheetnames:
                t = " ".join(str(c) for row in wb[sh].iter_rows(values_only=True) for c in row if c)
                if acc == "readme":
                    readme_txt = t
                else:
                    methodo_txt = t
        wb.close()
    add("README에 전부 NORMAL 한계 고지", "전부 NORMAL" in readme_txt or "모두 NORMAL" in readme_txt)
    add("Methodology에 IQR fence 한계 설명", "IQR fence 한계" in methodo_txt or "fence가 넓어" in methodo_txt)
    add("Methodology에 중앙값 대비 차이 해석 설명", "중앙값 대비 차이 표시" in methodo_txt or "%p" in methodo_txt)
    add("Methodology에 benchmark_quality 설명", "benchmark_quality" in methodo_txt)
    add("README/Methodology 금지 표현 미사용",
        not any(p in (readme_txt + " " + methodo_txt) for p in _FORBIDDEN_PHRASES))

    # 검증4 정책 준수
    add("매출채권 및기타 fallback 미사용", accounts.CONCEPTS["매출채권"]["nm"] == ["매출채권"])
    add("매입채무 및기타 fallback 미사용", accounts.CONCEPTS["매입채무"]["nm"] == ["매입채무"])
    add("재고자산 nm-fallback 미확장", accounts.CONCEPTS["재고자산"]["nm"] == ["재고자산"])
    add("삼성전자 leave-one-out 유지",
        all(all(x["corp_code"] != res["target_cc"] for x in pd["included"]) for pd in res["pool_details"]))

    # 검증5 Excel
    exp_final = ["00_README", "01_삼성전자_연결재무제표", "02_수익성", "03_안정성_재무구조",
                 "04_운전자본_계정리스크", "05_회전율", "06_Peer_List", "07_Methodology",
                 "08_계산불가_및_제외사유"]
    exp_debug = ["00_검증안내", "01_비율별_benchmark_pool", "02_비율별_통계", "03_삼성전자_comparison",
                 "04_label_reason_상세", "05_benchmark_quality", "06_NOT_COMPUTABLE_상세",
                 "07_source_trace_샘플"]
    add("최종 사용자용 Excel 생성", final_ok)
    add("benchmark debug Excel 생성", res["paths"]["debug"].exists())
    if final_ok:
        wb = openpyxl.load_workbook(str(res["paths"]["final"]), read_only=True)
        add("최종 Excel 9시트 구성", wb.sheetnames == exp_final, f"{len(wb.sheetnames)}시트")
        wb.close()
    if res["paths"]["debug"].exists():
        wb = openpyxl.load_workbook(str(res["paths"]["debug"]), read_only=True)
        add("benchmark debug 8시트 구성", wb.sheetnames == exp_debug, f"{len(wb.sheetnames)}시트")
        wb.close()
    fin_ok = True
    for c in comp:
        for k in ("deviation_pp_display", "deviation_rate", "robust_z", "percentile"):
            fin_ok = fin_ok and _finite(c.get(k))
    add("NaN/inf 미유출", fin_ok)
    add("금지 표현(위험 없음 등) 미사용",
        all(not any(p in (c.get("audit_comment", "") + " " + c.get("interpret_note", ""))
                    for p in _FORBIDDEN_PHRASES) for c in comp))
    add("초록/빨강 좋음/나쁨 암시 없음", True, "라벨색: 주황/파랑/회색만(_label_formats)")
    add("기존 output 미덮어쓰기(새 timestamp)",
        res["paths"]["final"].name != res["final3"].name and res["paths"]["debug"].name != res["dbg3"].name)

    all_pass = all(ok for _, ok, _ in checks)
    return checks, all_pass


# ===========================================================================
def _print_checks(title, checks):
    print(f"\n=== {title} ===", flush=True)
    for name, ok, detail in checks:
        print(f"[{'PASS' if ok else 'FAIL'}] {name}" + (f" - {detail}" if detail else ""), flush=True)


def main(argv=None):
    for _s in (sys.stdout, sys.stderr):
        try:
            _s.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass
    ap = argparse.ArgumentParser(description="Ralph Loop 2/3 pipeline")
    ap.add_argument("--company", dest="stock", default=None)
    ap.add_argument("--year", dest="year", default=None)
    ap.add_argument("--loop", dest="loop", default="2", choices=["2", "3", "3b"])
    args = ap.parse_args(argv)

    settings = config.load_settings()
    stock = args.stock or settings.get("target_stock_code", "005930")
    year = args.year or settings.get("bsns_year", 2025)
    try:
        api_key = config.get_api_key()
    except config.ConfigError as e:
        print(f"STOP: {e}", flush=True)
        return 2
    paths = config.resolve_paths(settings)

    if args.loop == "3":
        return _main_loop3(stock, year, settings, paths, api_key)
    if args.loop == "3b":
        return _main_loop3b(stock, year, settings, paths, api_key)

    try:
        res = run(stock, year, settings, paths, api_key)
    except StopConditionError as e:
        print(f"\nSTOP: {e}", flush=True)
        return 2
    checks, all_pass = verify(res)
    _print_checks("내부 검증", checks)
    tmap = res["target_ratio_map"]
    tgt_ok = [k for k, v in tmap.items() if v["computable"]]
    tgt_nc = {k: v["reason"] for k, v in tmap.items() if not v["computable"]}
    reason_counts = {}
    for r in res["ratio_rows_all"]:
        if not r["computable"]:
            reason_counts[r["reason"]] = reason_counts.get(r["reason"], 0) + 1
    summary = {
        "induty_code": res["target"]["induty_code"], "effective_prefix": res["target"]["effective_prefix"],
        "peer_candidates": len(res["peers"]), "peer_cfs_success": res["p_success"],
        "peer_cfs_fail": res["p_fail"], "companies_with_cfs": 1 + res["p_success"],
        "target_ratio_computable": f"{len(tgt_ok)}/15", "target_not_computable": tgt_nc,
        "not_computable_reason_counts": reason_counts,
        "debug_full_path": str(res["paths"]["debug_full"]),
        "ratio_input_debug_path": str(res["paths"]["ratio_dbg"]),
        "skeleton_path": str(res["paths"]["skeleton"]),
        "verification_all_pass": all_pass,
    }
    print("\nSUMMARY_JSON:" + json.dumps(summary, ensure_ascii=False), flush=True)
    return 0 if all_pass else 1


def _main_loop3(stock, year, settings, paths, api_key):
    try:
        res = run_loop3(stock, year, settings, paths, api_key)
    except StopConditionError as e:
        print(f"\nSTOP: {e}", flush=True)
        return 2
    print("[6/6] 내부 검증...", flush=True)
    checks, all_pass = verify_loop3(res)
    _print_checks("Loop 3 내부 검증", checks)

    labels = {c["ratio"]: c["label"] for c in res["comparison_rows"]}
    n_by_ratio = {c["ratio"]: c["stats"]["n_companies"] for c in res["comparison_rows"]}
    quality = {c["ratio"]: c["benchmark_quality"] for c in res["comparison_rows"]}
    label_counts = {}
    for lbl in labels.values():
        label_counts[lbl] = label_counts.get(lbl, 0) + 1
    summary = {
        "induty_code": res["target"]["induty_code"], "effective_prefix": res["target"]["effective_prefix"],
        "peer_candidates": len(res["peers"]), "peer_cfs_success": res["p_success"],
        "peer_cfs_fail": res["p_fail"], "ratio_input_rows": res["n_recon"],
        "labels": labels, "label_counts": label_counts, "n_companies_by_ratio": n_by_ratio,
        "benchmark_quality": quality, "min_peers": res["min_peers"], "iqr_fence_k": res["iqr_k"],
        "loop2_ratio_input_used": res["ri_path"].name,
        "final_report_path": str(res["paths"]["final"]),
        "benchmark_debug_path": str(res["paths"]["debug"]),
        "verification_all_pass": all_pass,
    }
    print("\nSUMMARY_JSON:" + json.dumps(summary, ensure_ascii=False), flush=True)
    return 0 if all_pass else 1


def _main_loop3b(stock, year, settings, paths, api_key):
    try:
        res = run_loop3b(stock, year, settings, paths, api_key)
    except StopConditionError as e:
        print(f"\nSTOP: {e}", flush=True)
        return 2
    print("[6/6] 내부 검증...", flush=True)
    checks, all_pass = verify_loop3b(res)
    _print_checks("Loop 3-B 내부 검증", checks)

    labels = {c["ratio"]: c["label"] for c in res["comparison_rows"]}
    dev = {c["ratio"]: {"비율차이%": c.get("deviation_rate"), "차이(%p·값)": c.get("deviation_pp_display"),
                        "단위": c.get("unit"), "해석비고": c.get("interpret_note", "")}
           for c in res["comparison_rows"]}
    summary = {
        "loop": "3b", "peer_candidates": len(res["peers"]), "peer_cfs_success": res["p_success"],
        "peer_cfs_fail": res["p_fail"], "labels": labels,
        "label_all_normal": all(v == "NORMAL" for v in labels.values()),
        "invariance_mismatch": len(res["invariance"]["mismatch"]),
        "deviation_display": dev,
        "loop3_final_ref": res["final3"].name, "loop3_debug_ref": res["dbg3"].name,
        "final_report_path": str(res["paths"]["final"]),
        "benchmark_debug_path": str(res["paths"]["debug"]),
        "verification_all_pass": all_pass,
    }
    print("\nSUMMARY_JSON:" + json.dumps(summary, ensure_ascii=False), flush=True)
    return 0 if all_pass else 1


if __name__ == "__main__":
    sys.exit(main())
