"""Ralph Loop 4 UI helpers — read-only Excel discovery/parsing + safe key handling.

분석 엔진(benchmark/compare/ratio/accounts)을 import·변경하지 않는다. 기존 산출물 Excel을
읽기 전용으로 열어 요약/표/해석문구/다운로드 bytes를 만든다. API key는 절대 반환/로그하지 않는다.
"""
from __future__ import annotations

import math
import re
from pathlib import Path

import openpyxl
import pandas as pd

# 그룹 탭 -> Excel 시트명
RATIO_TAB_SHEETS = {
    "수익성": "02_수익성",
    "안정성/재무구조": "03_안정성_재무구조",
    "운전자본/계정리스크": "04_운전자본_계정리스크",
    "회전율": "05_회전율",
}
RATIO_SHEETS = list(RATIO_TAB_SHEETS.values())

# --------------------------------------------------------------------------
# 파일 탐지 (읽기 전용, .tmp·중간 실패 파일 무시)
# --------------------------------------------------------------------------
# Loop 11: 삼성 전용 최신탐지 함수(_find_latest/find_latest_report/find_latest_debug)와
# resolve_report_pair 는 제거했다(회사 하드코딩 pattern 포함, app 미호출). app은 회사 일반
# 함수(list_report_candidates + resolve_pair_for_report)를 쓰며, 동등 최신탐지·pair 매칭
# 커버리지는 test_ui_dartlens.py 에 있다.
def year_from_name(path) -> str:
    m = re.search(r"_(\d{4})_\d{8}_\d{6}", Path(path).name)
    return m.group(1) if m else ""


def timestamp_from_name(path) -> str:
    """파일명에서 YYYYMMDD_HHMMSS timestamp 추출(없으면 '')."""
    m = re.search(r"_(\d{8}_\d{6})\.xlsx$", Path(path).name)
    return m.group(1) if m else ""


def company_from_report_name(path) -> str:
    """리포트 파일명에서 회사 접두(예: '현대자동차')를 추출(없으면 '')."""
    m = re.match(r"^(?P<company>.+?)_산업대비_이상징후_리포트_\d{4}_\d{8}_\d{6}\.xlsx$", Path(path).name)
    return m.group("company") if m else ""


def find_debug_for_report(output_dir, report_path, year=2025):
    """report와 **동일 timestamp**의 benchmark_debug 경로(없으면 None). .tmp 무시.

    두 명명 규칙을 모두 시도한다(읽기 전용):
      - multi-target(Loop 5/6): benchmark_debug_{회사}_{year}_{ts}.xlsx
      - run_loop3b(삼성 MVP):   benchmark_debug_{year}_{ts}.xlsx
    """
    ts = timestamp_from_name(report_path)
    if not ts:
        return None
    out = Path(output_dir)
    company = company_from_report_name(report_path)
    candidates = []
    if company:
        candidates.append(out / f"benchmark_debug_{company}_{year}_{ts}.xlsx")
    candidates.append(out / f"benchmark_debug_{year}_{ts}.xlsx")
    for cand in candidates:
        if cand.exists() and cand.suffix == ".xlsx":
            return cand
    return None


def pair_status_text(pair: dict):
    """(level, 메시지) 반환. level ∈ {'ok','warn'}."""
    status = pair.get("status")
    ts = pair.get("timestamp", "")
    if status == "matched":
        return "ok", f"최종 리포트와 Debug 파일 timestamp 일치 ({ts})"
    if status == "debug_missing":
        return "warn", (f"최신 리포트({ts})에 대응하는 동일 timestamp benchmark_debug 파일이 없습니다. "
                        "다른 timestamp debug를 조용히 내려받지 않도록 Debug 다운로드를 비활성화합니다.")
    return "warn", "표시할 최종 리포트를 찾지 못했습니다."


# --------------------------------------------------------------------------
# Excel 읽기 (읽기 전용, header 자동 탐지, NaN/inf 안전 처리)
# --------------------------------------------------------------------------
def list_sheets(path) -> list[str]:
    wb = openpyxl.load_workbook(str(path), read_only=True)
    names = list(wb.sheetnames)
    wb.close()
    return names


def _sanitize(df: pd.DataFrame) -> pd.DataFrame:
    return df.replace([math.inf, -math.inf], pd.NA)


def sheet_to_df(path, sheet, header_contains=None, max_scan=6) -> pd.DataFrame:
    """시트를 DataFrame으로 읽는다. header_contains가 있으면 그 값을 포함한 행을 헤더로 탐지
    (상단 병합 안내행 skip). 읽기 전용, 값 미수정."""
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[sheet]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        return pd.DataFrame()
    hdr = 0
    if header_contains is not None:
        for i, r in enumerate(rows[:max_scan]):
            if r and any(c == header_contains for c in r):
                hdr = i
                break
    header = ["" if c is None else str(c) for c in rows[hdr]]
    data = rows[hdr + 1:]
    df = pd.DataFrame(data, columns=header)
    df = df.dropna(how="all").reset_index(drop=True)
    return _sanitize(df)


def ratio_sheet_df(path, sheet) -> pd.DataFrame:
    return sheet_to_df(path, sheet, header_contains="비율명")


def combined_ratio_df(path) -> pd.DataFrame:
    """02~05 비율 시트를 합친 15행 DataFrame."""
    frames = []
    for sheet in RATIO_SHEETS:
        try:
            frames.append(ratio_sheet_df(path, sheet))
        except Exception:
            continue
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


# --------------------------------------------------------------------------
# 요약/해석 추출
# --------------------------------------------------------------------------
def _first_num(series):
    for v in series:
        if v is None or (isinstance(v, float) and math.isnan(v)):
            continue
        try:
            return int(float(v))
        except (TypeError, ValueError):
            continue
    return None


def target_company_name(final_path) -> str:
    """06_Peer_List에서 대상여부=='대상' 회사명. 못 읽으면 빈 문자열(미확인) 반환.

    Loop 11: 특정회사(삼성전자) fallback 제거 — 비삼성 리포트에서 대상명을 못 읽었을 때
    '삼성전자(주)'로 잘못 표시되던 latent bug를 없앤다. 빈 값은 호출부에서 '-'로 표기한다.
    """
    try:
        df = sheet_to_df(final_path, "06_Peer_List", header_contains="기업코드")
        if "대상여부" in df.columns and "기업명" in df.columns:
            hit = df[df["대상여부"] == "대상"]
            if len(hit):
                return str(hit.iloc[0]["기업명"])
    except Exception:
        pass
    return ""


def extract_summary(final_path) -> dict:
    """요약 카드용 값. 데이터 없으면 안전한 부분값(None)로 반환."""
    out = {"company": target_company_name(final_path), "year": year_from_name(final_path),
           "peer_candidates": None, "cfs_success": None, "cfs_fail": None,
           "computable_count": None, "total_ratios": None, "label_counts": {}}
    df = combined_ratio_df(final_path)
    if df.empty:
        return out
    out["total_ratios"] = len(df)
    # peer/CFS 수: 06_Peer_List 우선(신 9열 리포트), 없으면 비율시트 컬럼(구 24열 리포트) fallback
    cand, succ, fail = _peer_counts_from_06(final_path)
    if cand is not None:
        out["peer_candidates"], out["cfs_success"], out["cfs_fail"] = cand, succ, fail
    else:
        if "peer 후보 수" in df.columns:
            out["peer_candidates"] = _first_num(df["peer 후보 수"])
        if "CFS 성공 peer 수" in df.columns:
            out["cfs_success"] = _first_num(df["CFS 성공 peer 수"])
        if "CFS 실패 peer 수" in df.columns:
            out["cfs_fail"] = _first_num(df["CFS 실패 peer 수"])
    jcol = _first_col(df, "상대판정", "판정")
    if jcol:
        labels = [str(x) for x in df[jcol].tolist() if x is not None and str(x) != "nan"]
        out["computable_count"] = sum(1 for l in labels if l != "계산 불가")
        counts = {}
        for l in labels:
            counts[l] = counts.get(l, 0) + 1
        out["label_counts"] = counts
    return out


# Loop 15: 비율 시트 컬럼 별칭(구 24열 리포트/신 9열 리포트 모두 읽기). 앞이 신, 뒤가 구.
def _first_col(df, *names):
    for n in names:
        if n in df.columns:
            return n
    return None


def _peer_counts_from_06(final_path):
    """06_Peer_List에서 peer 후보/CFS 성공·실패 수(신 9열 리포트는 비율시트에 없어 여기서 조달)."""
    try:
        df = sheet_to_df(final_path, "06_Peer_List", header_contains="기업코드")
    except Exception:
        return None, None, None
    if "대상여부" not in df.columns or "CFS수집상태" not in df.columns:
        return None, None, None
    peers = df[df["대상여부"] == "peer"]
    cand = len(peers)
    succ = sum(1 for x in peers["CFS수집상태"].tolist() if str(x) == "성공")
    return cand, succ, cand - succ


def _get(df, ratio_name, col):
    if "비율명" not in df.columns or col not in df.columns:
        return None
    hit = df[df["비율명"] == ratio_name]
    if not len(hit):
        return None
    v = hit.iloc[0][col]
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return None
    return v


def build_interpretation(final_path) -> list[str]:
    """실제 label/percentile/benchmark_quality에서 핵심 해석 문장 생성. 데이터 없으면 기본 안내."""
    df = combined_ratio_df(final_path)
    jcol = _first_col(df, "상대판정", "판정") if not df.empty else None
    if df.empty or jcol is None:
        return ["분석 결과 데이터를 읽지 못했습니다. output 폴더의 최신 리포트를 확인하세요."]
    pcol = _first_col(df, "산업 내 위치(percentile)", "percentile")
    qcol = _first_col(df, "benchmark_quality", "신뢰도(peer·품질)")
    acol = _first_col(df, "절대판정(red flag)")
    lines = []
    # Loop 24: 결과 화면 상단 티어와 동일한 상호배타 6분류 → 합이 항상 총 비율 수와 일치(누락 표시 방지).
    # 우선순위: 절대판정 경고 > 절대판정 주의 > 상대판정 업종 높음/낮음 > 계산 제한(peer·분포 부족·계산
    # 불가) > 정상. 표시 전용 카운트 — 엔진이 산출한 label/절대판정을 '읽어' 분류만 한다(값·판정 불변, INV-7).
    _INFO_LABELS = {"peer 부족", "분포 부족", "계산 불가"}
    cnt = {"normal": 0, "warn": 0, "caution": 0, "high": 0, "low": 0, "info": 0}
    for _, r in df.iterrows():
        absr = str(r.get(acol)).strip() if acol else ""
        rel = str(r.get(jcol)).strip()
        if absr == "경고":
            cnt["warn"] += 1
        elif absr == "주의":
            cnt["caution"] += 1
        elif rel == "산업 대비 높음":
            cnt["high"] += 1
        elif rel == "산업 대비 낮음":
            cnt["low"] += 1
        elif rel in _INFO_LABELS:
            cnt["info"] += 1
        else:
            cnt["normal"] += 1
    n = sum(cnt.values())
    if n and cnt["normal"] == n:
        lines.append(f"총 {n}개 비율 모두 현재 peer universe·IQR 기준상 정상 범위입니다(이상치로 분류되지 않음).")
    else:
        lines.append(
            f"총 {n}개 비율 중 정상 {cnt['normal']} · 절대 경고 {cnt['warn']} · 절대 주의 {cnt['caution']} · "
            f"계산 제한 {cnt['info']} (업종 대비 높음 {cnt['high']} · 낮음 {cnt['low']})."
        )
    # 절대판정(red flag) — 상대판정과 별개로 절대 기준선 경고/주의를 표면화(위험 확정 아님)
    if acol:
        warns = [str(r["비율명"]) for _, r in df.iterrows() if str(r.get(acol)) in ("경고", "주의")]
        if warns:
            lines.append(f"절대판정 red flag: {', '.join(warns[:6])}에서 경고/주의 신호가 있습니다"
                         "(검토가 필요한 점검 신호이며 위험 확정이 아닙니다).")
    # 영업이익률 상/하위권
    op_label = _get(df, "영업이익률", jcol)
    op_pct = _get(df, "영업이익률", pcol) if pcol else None
    if op_label is not None and op_pct is not None:
        try:
            p = float(op_pct)
            if p >= 90:
                lines.append(f"영업이익률은 산업 내 상위권(percentile ≈ {p:.0f})이나 IQR 이상치 기준은 초과하지 않았습니다.")
            elif p <= 10:
                lines.append(f"영업이익률은 산업 내 하위권(percentile ≈ {p:.0f})이나 IQR 이상치 기준은 초과하지 않았습니다.")
        except (TypeError, ValueError):
            pass
    # Loop 24: 신뢰도 제한 비율 안내 — 영어 코드(WEAK/LIMITED) 제거, 이유를 한글로 명시.
    # benchmark_quality 정의상 제한 사유는 '계산 가능 peer 표본이 작거나(n<2*min_peers) 계정
    # 커버리지(n/CFS성공)가 낮음'이며 고정 peer 임계값이 아니다(coverage 부족은 peer가 많아도 발생 —
    # 예: peer 22·26에서도 WEAK). 따라서 특정 개수를 지어내지 않고 실제 제한 비율만 동적으로 나열한다.
    if qcol and "비율명" in df.columns:
        limited = [str(r["비율명"]) for _, r in df.iterrows()
                   if str(r.get(qcol)).upper().startswith(("WEAK", "LIMITED"))]
        if limited:
            lines.append(" · ".join(limited[:4])
                         + " 항목은 비교 가능한 동종업체(peer)가 적어 산업 비교의 신뢰도가 제한적입니다.")
    return lines


# --------------------------------------------------------------------------
# 다운로드
# --------------------------------------------------------------------------
def prepare_download(path):
    """(파일명, bytes) 반환. 읽기 전용."""
    p = Path(path)
    return p.name, p.read_bytes()


# --------------------------------------------------------------------------
# API key 안전 처리 (절대 key 값을 반환/로그하지 않음)
# --------------------------------------------------------------------------
def env_key_available() -> bool:
    """.env/환경에 OPENDART_API_KEY가 있으면 True(값은 반환하지 않음)."""
    try:
        from . import config
        config.get_api_key()
        return True
    except Exception:
        return False


def mask_key(key: str) -> str:
    """키 표시용 마스크. 실제 키 문자는 노출하지 않는다."""
    if not key:
        return "미설정"
    return "설정됨(****)"


def key_status_text(sidebar_key: str) -> str:
    """화면 표시용 키 상태 문자열(값 미노출)."""
    if sidebar_key:
        return "입력 키 사용(세션 한정, 저장 안 함)"
    if env_key_available():
        return ".env 키 사용 가능"
    return "키 미설정 — .env 또는 sidebar 입력 필요"


# --------------------------------------------------------------------------
# Ralph Loop 7-2: 표시(display) layer 전용 — 원본 status/label 데이터는 변경하지 않는다.
# --------------------------------------------------------------------------
# 개발자용 raw status/label → 사용자 표시 문구. debug/summary 원본은 그대로 두고 표시할 때만 변환.
STATUS_DISPLAY = {
    "PASS": "분석 완료",
    "PASS_WITH_WARNINGS": "분석 완료 · 판정 제한",
    "FAIL": "분석 실패",
    "INSUFFICIENT_PEERS": "표본 제한",
    "NOT_COMPUTABLE": "계산 불가",
}

SPARSE_SHEET_NAME = "09_제한적_peer_비교"


def status_display(raw) -> str:
    """raw status/label 문자열을 사용자 표시 문구로 변환(매핑 없으면 원문 유지). 표시 전용."""
    if raw is None:
        return ""
    return STATUS_DISPLAY.get(str(raw).strip(), str(raw))


def has_sparse_sheet(path) -> bool:
    """리포트에 09_제한적_peer_비교(sparse peer 참고 비교) 시트가 있으면 True(읽기 전용)."""
    try:
        return SPARSE_SHEET_NAME in list_sheets(path)
    except Exception:
        return False


# raw status 토큰 표기 legend 순서(방법론/보조 영역에서 원시값→표시 대응 안내용).
STATUS_LEGEND_TOKENS = ["PASS", "PASS_WITH_WARNINGS", "FAIL", "INSUFFICIENT_PEERS", "NOT_COMPUTABLE"]

# run status 도출 시 '판정 제한'으로 보는 리포트 라벨(엔진 status를 재계산하지 않는 표시용 요약).
# peer 부족(INSUFFICIENT_PEERS)만 사용 — multi_target_runner의 PASS_WITH_WARNINGS(peer<min_peers 등)와
# 일치하도록, 소수 '계산 불가'(NOT_COMPUTABLE)만 있는 리포트는 PASS로 둔다(요약 xlsx와 표시 일관성).
_LIMITED_LABELS = {"peer 부족"}


def report_status_token(final_path) -> str:
    """리포트에서 사용자 표시용 run status 토큰을 **읽기 전용** 도출(원본·엔진 status 재계산 아님).

    sparse peer 시트가 있거나 peer 부족(표본 제한) 판정이 있으면 PASS_WITH_WARNINGS, 아니면 PASS.
    소수 '계산 불가'만 있는 경우는 PASS(요약 status와 일관). status_display()로 한글 표시한다.
    """
    try:
        if has_sparse_sheet(final_path):
            return "PASS_WITH_WARNINGS"
        df = combined_ratio_df(final_path)
        cols = getattr(df, "columns", [])
        if "판정" in cols:
            labels = {str(x) for x in df["판정"].tolist()}
            if labels & _LIMITED_LABELS:
                return "PASS_WITH_WARNINGS"
    except Exception:
        pass
    return "PASS"


# --------------------------------------------------------------------------
# Loop 7-2A: 최근 결과 후보 탐지(삼성 전용이 아니라 Loop 5/6 산출물 전체, 읽기 전용)
# --------------------------------------------------------------------------
_REPORT_RE = re.compile(
    r"^(?P<company>.+?)_산업대비_이상징후_리포트_(?P<year>\d{4})_(?P<ts>\d{8}_\d{6})\.xlsx$")


def list_report_candidates(output_dir, limit=20) -> list[dict]:
    """output의 '..._산업대비_이상징후_리포트_{year}_{ts}.xlsx'를 **최신 timestamp 우선**으로 후보화.

    회사 접두를 가리지 않으므로 삼성전자뿐 아니라 CJ제일제당·한화솔루션·현대자동차·대한항공 등
    Loop 5/6 산출물도 후보가 된다. 파일명만으로 판별(읽기 전용, .tmp 무시). 각 후보:
      {path, company, year, timestamp, filename}
    """
    out = Path(output_dir)
    if not out.exists():
        return []
    cands = []
    for p in out.glob("*_산업대비_이상징후_리포트_*.xlsx"):
        if p.suffix != ".xlsx" or p.name.endswith(".tmp"):
            continue
        m = _REPORT_RE.match(p.name)
        if not m:
            continue
        cands.append({"path": p, "company": m.group("company"), "year": m.group("year"),
                      "timestamp": m.group("ts"), "filename": p.name})
    cands.sort(key=lambda d: d["timestamp"], reverse=True)   # 사전식 timestamp == 시간순
    return cands[:limit]


def candidate_label(cand: dict) -> str:
    """selectbox 표시용 후보 라벨: '회사 · 연도 · timestamp'."""
    return f"{cand.get('company', '')} · {cand.get('year', '')} · {cand.get('timestamp', '')}"


def resolve_pair_for_report(output_dir, report_path, year=None) -> dict:
    """선택한 특정 report에 대해 동일 timestamp debug를 매칭(resolve_report_pair의 후보 지정판).

    반환 형태는 resolve_report_pair와 동일: {report, debug, pair_ok, status, timestamp}.
    """
    report = Path(report_path)
    if not report.exists():
        return {"report": None, "debug": None, "pair_ok": False, "status": "no_report", "timestamp": ""}
    if year is None:
        y = year_from_name(report)
        year = int(y) if y.isdigit() else 2025
    ts = timestamp_from_name(report)
    debug = find_debug_for_report(output_dir, report, year)
    if debug is not None:
        return {"report": report, "debug": debug, "pair_ok": True, "status": "matched", "timestamp": ts}
    return {"report": report, "debug": None, "pair_ok": False, "status": "debug_missing", "timestamp": ts}
